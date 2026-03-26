from __future__ import annotations

import io
import json
import math
import mimetypes
import os
import re
import shutil
import time
import uuid
from copy import copy as copy_style
from functools import lru_cache
from datetime import date, datetime, time as dt_time, timedelta
from pathlib import Path
from random import choice, sample
from collections import defaultdict
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse

from flask import Flask, Response, jsonify, redirect, render_template, request, send_file, stream_with_context, url_for
from dotenv import load_dotenv
import requests
from sqlalchemy import Boolean, Date, DateTime, ForeignKey, Integer, String, Text, UniqueConstraint, create_engine, func, select, update, delete
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column, relationship, sessionmaker
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


load_dotenv()


app = Flask(__name__)

ASSET_VERSION = os.getenv("ASSET_VERSION", "20260318")
app.jinja_env.globals["ASSET_VERSION"] = ASSET_VERSION

BASE_DIR = Path(app.root_path)
SCHEDULE_TEMPLATE_FILENAME = BASE_DIR / "ScheduleTemplate.xlsx"
SCHEDULE_TEMPLATE_ARCHIVE_DIR = BASE_DIR / "old_schedule_templates"
SCHEDULE_TEMPLATE_ALLOWED_SUFFIXES = {".xlsx"}


TWILIO_MESSAGES_URL_TEMPLATE = "https://api.twilio.com/2010-04-01/Accounts/{account_sid}/Messages.json"
WHATSAPP_MAX_IMAGE_BYTES = 8 * 1024 * 1024  # 8 MB
WHATSAPP_TIMEOUT = 15  # seconds
WHATSAPP_MEDIA_RELATIVE = Path("static/whatsapp-media")
WHATSAPP_MEDIA_DIR = Path(app.root_path) / WHATSAPP_MEDIA_RELATIVE
WHATSAPP_MEDIA_MAX_AGE_SECONDS = 24 * 60 * 60


class TwilioWhatsAppError(RuntimeError):
    """Raised when the Twilio WhatsApp API reports an error."""


def _normalize_whatsapp_number(raw: str) -> str:
    value = (raw or "").strip()
    if not value:
        return ""
    if not value.startswith("whatsapp:"):
        value = f"whatsapp:{value}"
    return value


def _twilio_config() -> Optional[dict[str, str]]:
    account_sid = (os.getenv("TWILIO_ACCOUNT_SID") or "").strip()
    auth_token = (os.getenv("TWILIO_AUTH_TOKEN") or "").strip()
    from_number = _normalize_whatsapp_number(os.getenv("TWILIO_WHATSAPP_FROM") or "whatsapp:+14155238886")
    to_number = _normalize_whatsapp_number(os.getenv("TWILIO_WHATSAPP_TO") or "")
    media_base_url = (os.getenv("WHATSAPP_MEDIA_BASE_URL") or "").strip().rstrip("/")
    if not account_sid or not auth_token or not to_number or not media_base_url:
        return None
    return {
        "account_sid": account_sid,
        "auth_token": auth_token,
        "from_number": from_number,
        "to_number": to_number,
        "media_base_url": media_base_url,
    }


def _ensure_whatsapp_media_dir() -> None:
    try:
        WHATSAPP_MEDIA_DIR.mkdir(parents=True, exist_ok=True)
    except Exception as exc:  # pragma: no cover - filesystem guard
        app.logger.warning("Unable to create WhatsApp media dir: %s", exc)


def _cleanup_old_whatsapp_media() -> None:
    if not WHATSAPP_MEDIA_DIR.exists():
        return
    cutoff = time.time() - WHATSAPP_MEDIA_MAX_AGE_SECONDS
    for path in WHATSAPP_MEDIA_DIR.iterdir():
        if not path.is_file():
            continue
        try:
            if path.stat().st_mtime < cutoff:
                path.unlink(missing_ok=True)  # type: ignore[arg-type]
        except Exception:
            continue


def _persist_whatsapp_media(file_bytes: bytes, extension: str) -> str:
    _ensure_whatsapp_media_dir()
    _cleanup_old_whatsapp_media()
    safe_ext = extension if extension.startswith(".") else f".{extension}" if extension else ".bin"
    filename = f"{int(time.time())}-{uuid.uuid4().hex}{safe_ext}"
    target_path = WHATSAPP_MEDIA_DIR / filename
    with open(target_path, "wb") as f:
        f.write(file_bytes)
    return filename


def _build_media_url(base_url: str, filename: str) -> str:
    normalized = base_url.rstrip("/") + "/"
    return urljoin(normalized, filename)


def _sanitize_redirect_target(raw: Optional[str]) -> str:
    if not raw:
        return ""
    value = raw.strip()
    if not value:
        return ""
    lowered = value.lower()
    if lowered.startswith("http://") or lowered.startswith("https://") or value.startswith("//"):
        return ""
    if not value.startswith("/"):
        return ""
    return value


def _redirect_with_template_status(target_url: str, status: str, message: str) -> Response:
    sanitized_status = status if status in {"success", "error"} else "error"
    trimmed_message = (message or "").strip()
    if len(trimmed_message) > 200:
        trimmed_message = trimmed_message[:200]
    parsed = list(urlparse(target_url or "/"))
    query_items = dict(parse_qsl(parsed[4], keep_blank_values=True))
    query_items["template_status"] = sanitized_status
    query_items["template_message"] = trimmed_message
    parsed[4] = urlencode(query_items)
    return redirect(urlunparse(parsed))


def _send_twilio_whatsapp_image(config: dict[str, str], *, media_url: str, caption: Optional[str]) -> None:
    url = TWILIO_MESSAGES_URL_TEMPLATE.format(account_sid=config["account_sid"])
    data = {
        "From": config["from_number"],
        "To": config["to_number"],
        "MediaUrl": media_url,
    }
    if caption:
        data["Body"] = caption[:1024]
    try:
        response = requests.post(
            url,
            data=data,
            auth=(config["account_sid"], config["auth_token"]),
            timeout=WHATSAPP_TIMEOUT,
        )
    except requests.RequestException as exc:
        raise TwilioWhatsAppError(f"Unable to reach Twilio ({exc})") from exc
    if response.status_code >= 400:
        message: Optional[str] = None
        try:
            payload = response.json()
            if isinstance(payload, dict):
                message = payload.get("message") or payload.get("error_message")
        except ValueError:
            message = None
        if not message:
            message = f"Twilio error ({response.status_code})"
        raise TwilioWhatsAppError(message)


def _post_discord_message(content: str, *, title: Optional[str] = None, color: Optional[int] = None) -> None:
    """Send an embed to the configured Discord webhook, if available."""
    url = os.getenv("DISCORD_WEBHOOK_URL")
    if not url or not content:
        return
    embed_payload = {
        "description": content,
    }
    if title:
        embed_payload["title"] = title
    if color is not None:
        embed_payload["color"] = color
    try:
        response = requests.post(url, json={"embeds": [embed_payload]}, timeout=5)
        if response.status_code >= 400:
            app.logger.warning(
                "Discord webhook returned %s: %s",
                response.status_code,
                response.text[:200],
            )
    except Exception as exc:  # pragma: no cover - defensive logging only
        app.logger.warning("Discord webhook error: %s", exc)


def _notify_schedule_change(employee: str, section: str, shift_date: date, value: str) -> None:
    message = (
        "Schedule update: "
        f"{employee} assigned to {value or 'Set'} on {shift_date.isoformat()} ({section})."
    )
    _post_discord_message(message, title="Schedule Updated", color=0x5865F2)


def _notify_timeoff_submission(name: str, role: str, start: date, end: date, approved: bool, vacation: bool) -> None:
    status = "approved" if approved else "pending approval"
    vacation_note = "vacation" if vacation else "time off"
    message = (
        "Time off request submitted: "
        f"{name} ({role}) {vacation_note} from {start.isoformat()} to {end.isoformat()} ({status})."
    )
    _post_discord_message(message, title="Time Off Submitted", color=0x57F287)


@app.after_request
def add_no_cache_headers(resp):
    # Ensure freshly generated schedules show updated coverage highlights immediately
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


# ---- SQLAlchemy setup ----
class Base(DeclarativeBase):
    pass


class Section(Base):
    __tablename__ = "sections"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    name: Mapped[str] = mapped_column(String, unique=True)
    required_per_day: Mapped[Optional[int]] = mapped_column(Integer, nullable=True)
    employees: Mapped[List["Employee"]] = relationship("Employee", back_populates="section")


class Employee(Base):
    __tablename__ = "employees"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    name: Mapped[str] = mapped_column(String)
    section_id: Mapped[int] = mapped_column(ForeignKey("sections.id"))
    section: Mapped[Section] = relationship("Section", back_populates="employees")
    # New editable fields
    availability: Mapped[Optional[str]] = mapped_column(String, nullable=True)  # freeform notes or JSON
    preferred_shift: Mapped[Optional[str]] = mapped_column(String, nullable=True)
    seniority: Mapped[Optional[int]] = mapped_column(Integer, nullable=True)
    preferred_shifts_per_week: Mapped[Optional[int]] = mapped_column(Integer, nullable=True)
    max_shifts_per_week: Mapped[Optional[int]] = mapped_column(Integer, nullable=True)
    sort_order: Mapped[Optional[int]] = mapped_column(Integer, nullable=True)
    first_name: Mapped[Optional[str]] = mapped_column(String, nullable=True)
    last_name: Mapped[Optional[str]] = mapped_column(String, nullable=True)


class EmployeeRole(Base):
    __tablename__ = "employee_roles"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    employee_id: Mapped[int] = mapped_column(ForeignKey("employees.id"))
    section_id: Mapped[int] = mapped_column(ForeignKey("sections.id"))


class EmployeeAvailability(Base):
    __tablename__ = "employee_availability"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    employee_id: Mapped[int] = mapped_column(ForeignKey("employees.id"))
    day_of_week: Mapped[int] = mapped_column(Integer)  # 0=Mon .. 6=Sun
    shift_label: Mapped[str] = mapped_column(String)   # matches one of role shift variants
    allowed: Mapped[bool] = mapped_column(Boolean, default=True)


class Week(Base):
    __tablename__ = "weeks"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    start_date: Mapped[date] = mapped_column(Date, unique=True)  # Thu start


class Assignment(Base):
    __tablename__ = "assignments"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    week_id: Mapped[int] = mapped_column(ForeignKey("weeks.id"))
    employee_id: Mapped[int] = mapped_column(ForeignKey("employees.id"))
    date: Mapped[date] = mapped_column(Date)
    value: Mapped[str] = mapped_column(String)  # shift label, e.g., "Set", "6AM–2PM"
    # Mark when the scheduler would have assigned a shift but skipped due to time off
    dismissed_timeoff: Mapped[bool] = mapped_column(Boolean, default=False)


class AircrewArrival(Base):
    __tablename__ = "aircrew_arrivals"
    __table_args__ = (
        UniqueConstraint("week_id", "carrier", "date", name="uniq_aircrew_week_carrier_date"),
    )

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    week_id: Mapped[int] = mapped_column(ForeignKey("weeks.id"))
    carrier: Mapped[str] = mapped_column(String)
    date: Mapped[date] = mapped_column(Date)
    times: Mapped[str] = mapped_column(String, default="")


class OccupancySnapshot(Base):
    __tablename__ = "occupancy_levels"
    __table_args__ = (
        UniqueConstraint("week_id", "date", name="uniq_occupancy_week_date"),
    )

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    week_id: Mapped[int] = mapped_column(ForeignKey("weeks.id"))
    date: Mapped[date] = mapped_column(Date)
    percentage: Mapped[Optional[int]] = mapped_column(Integer, nullable=True)


class TimeOff(Base):
    __tablename__ = "timeoff"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    name: Mapped[str] = mapped_column(String)
    role: Mapped[str] = mapped_column(String)
    from_date: Mapped[date] = mapped_column(Date)
    to_date: Mapped[date] = mapped_column(Date)
    approved: Mapped[bool] = mapped_column(Boolean, default=False)
    vacation: Mapped[bool] = mapped_column(Boolean, default=False)


class ScheduleTemplate(Base):
    __tablename__ = "schedule_templates"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    slot: Mapped[int] = mapped_column(Integer, unique=True)
    name: Mapped[str] = mapped_column(String, nullable=False)
    payload: Mapped[str] = mapped_column(Text, nullable=False)
    saved_week_start: Mapped[Optional[date]] = mapped_column(Date, nullable=True)
    updated_at: Mapped[datetime] = mapped_column(
        DateTime, nullable=False, default=datetime.utcnow, onupdate=datetime.utcnow
    )


engine = create_engine("sqlite:///schedule.db", future=True)
SessionLocal = sessionmaker(bind=engine, future=True, expire_on_commit=False)


# ---- Helpers and constants ----
def format_employee_name(first_name: str, last_name: Optional[str]) -> str:
    """Build a display string from first/last name parts."""
    parts = [(first_name or "").strip()]
    if last_name:
        parts.append(last_name.strip())
    return " ".join(p for p in parts if p)


def employee_by_role(session: Session, *, name: str, role: str) -> Optional[Employee]:
    """Fetch an employee by name constrained to a specific role/section."""
    stmt = (
        select(Employee)
        .join(Section)
        .where(Employee.name == name)
    )
    if role:
        stmt = stmt.where(Section.name == role)
    return session.scalar(stmt)


def daterange(start: date, days: int) -> List[date]:
    return [start + timedelta(days=i) for i in range(days)]


def week_dates(start: date) -> List[dict]:
    result = []
    for d in daterange(start, 7):
        result.append(
            {
                "key": d.isoformat(),
                "label_short": d.strftime("%a %m/%d"),
                "label_long": d.strftime("%a %b %d, %Y"),
                "label_md": f"{d.month}/{d.day}",
            }
        )
    return result


def format_week_label(start: date) -> str:
    end = start + timedelta(days=6)
    return f"{start.strftime('%a %b %d, %Y')} – {end.strftime('%a %b %d, %Y')}"


FOUR_WEEK_BASELINE = date(2025, 9, 18)


def four_week_period_bounds(week_start: date) -> Tuple[date, date]:
    """Return the 4-week period (start, end) anchored to the baseline week."""
    if not week_start:
        return FOUR_WEEK_BASELINE, FOUR_WEEK_BASELINE + timedelta(days=27)
    delta_days = (week_start - FOUR_WEEK_BASELINE).days
    if delta_days >= 0:
        idx = delta_days // 28
    else:
        idx = -((-delta_days - 1) // 28) - 1
    period_start = FOUR_WEEK_BASELINE + timedelta(days=idx * 28)
    return period_start, period_start + timedelta(days=27)


def format_four_week_label(start: date, end: date) -> str:
    """Pretty label for 4-week period headers."""
    start_month = start.strftime("%B")
    end_month = end.strftime("%B")
    if start.year == end.year:
        if start.month == end.month:
            return f"{start_month} {start.day} – {end.day}, {start.year}"
        return f"{start_month} {start.day} – {end_month} {end.day}, {start.year}"
    return f"{start_month} {start.day}, {start.year} – {end_month} {end.day}, {end.year}"


def week_start_for_date(target: date) -> date:
    """Find the stored week start corresponding to a calendar date."""
    delta_days = (target - FOUR_WEEK_BASELINE).days
    weeks_offset = delta_days // 7
    return FOUR_WEEK_BASELINE + timedelta(days=weeks_offset * 7)


TIME_OFF_LABEL = "TIME OFF"
REQ_VAC_LABEL = "REQ VAC"
SHUTTLE_COMBO_LABEL = "10:30am - 6:30pm (c)"
SUGGESTED_CREW_REGEX = re.compile(r"^\s*\d{1,2}:\d{2}(?:am|pm)\s*-\s*\d{1,2}:\d{2}(?:am|pm)\s*$", re.IGNORECASE)
TIME_OFF_VALUES = {TIME_OFF_LABEL, REQ_VAC_LABEL}
NEUTRAL_ASSIGNMENT_VALUES = {"Set"} | TIME_OFF_VALUES
TEMPLATE_SLOT_COUNT = 3
OCCUPANCY_DATE_RE = re.compile(r"\b(\d{2})-(\d{2})-(\d{2})\b")
OCCUPANCY_PERCENT_RE = re.compile(r"(\d{1,3}(?:\.\d+)?)\s*%")


def is_suggested_crew_label(value: Optional[str]) -> bool:
    if not value:
        return False
    return SUGGESTED_CREW_REGEX.match(value.strip()) is not None


UNDO_DELETE_SECONDS = 20
_pending_period_undos: Dict[str, dict] = {}


def _prune_expired_period_undos() -> None:
    """Drop any undo tokens that have passed their lifetime."""
    now = time.time()
    expired = [tok for tok, data in _pending_period_undos.items() if data.get("expires", 0) <= now]
    for tok in expired:
        _pending_period_undos.pop(tok, None)

# Seniority order for Front Desk manager-on-duty selection
SENIORITY_ORDER = [
    "Cindy", "KC", "Ryan", "Emilyn", "Christian", "Troy",
    "Brian", "Tristan", "Terry", "Jordan", "Abdi", "Sato"
]

BREAKFAST_SHIFTS = [
    "Set",
    TIME_OFF_LABEL,
    REQ_VAC_LABEL,
    "5AM–12PM",
    "6AM–12PM",
    "7AM–12PM",
]

# Front Desk: three variants (AM, PM, Audit), each has two staggered times
FRONT_DESK_SHIFTS = [
    "Set",
    TIME_OFF_LABEL,
    REQ_VAC_LABEL,
    "AM (6:00AM–2:00PM)",
    "AM (6:15AM–2:15PM)",
    "PM (2:00PM–10:00PM)",
    "PM (2:15PM–10:15PM)",
    "Audit (10:00PM–6:00AM)",
    "Audit (10:15PM–6:15AM)",
]

SHUTTLE_CREW_SHIFTS = [
    "Crew (5:45PM–1:45AM)",
    "Crew (8:00PM–12:00AM)",
    "Crew (9:00PM–1:00AM)",
]

# Shuttle: fixed variants plus crew windows
SHUTTLE_SHIFTS = [
    "Set",
    TIME_OFF_LABEL,
    REQ_VAC_LABEL,
    "AM (3:30AM–11:30AM)",
    "Midday (10:30AM–6:30PM)",
    SHUTTLE_COMBO_LABEL,
    "PM (5:30PM–1:30AM)",
] + SHUTTLE_CREW_SHIFTS

DEFAULT_CREW_SHIFT = SHUTTLE_CREW_SHIFTS[0]
CREW_EXCEL_FILL = PatternFill(start_color="FFFFB347", end_color="FFFFB347", fill_type="solid")

MAINTENANCE_SHIFTS = [
    "Set",
    TIME_OFF_LABEL,
    REQ_VAC_LABEL,
    "8AM–4:30PM",
]

SECTION_SHIFT_MAP = {
    "Breakfast Bar": BREAKFAST_SHIFTS,
    "Front Desk": FRONT_DESK_SHIFTS,
    "Shuttle": SHUTTLE_SHIFTS,
    "Maintenance": MAINTENANCE_SHIFTS,
}

SECTION_DISPLAY_ORDER = ["Breakfast Bar", "Front Desk", "Shuttle", "Maintenance"]

AIRCREW_CARRIERS = ("Aeromexico", "Skywest")
AIRCREW_CARRIER_ALIAS_MAP = {
    "AEROMEXICO": "Aeromexico",
    "AEROMEXIC": "Aeromexico",
    "AEROMEXICOAIRLINES": "Aeromexico",
    "AEROMEXICOAIR": "Aeromexico",
    "AEROMEXICOAIRLINE": "Aeromexico",
    "AEROMEXICOARRIVALS": "Aeromexico",
    "AEROMEXICOAIRPORT": "Aeromexico",
    "AEROMEXICOAEROMEXICO": "Aeromexico",
    "AEROMEXICOFLIGHT": "Aeromexico",
    "SKYWEST": "Skywest",
    "SKYWESTAIRLINES": "Skywest",
}
AIRCREW_HEADER_SKIP = {"day", "days", "weekday", "week", "notes", "note", "flight", "flight#", "flight #", "flight number"}
EXCEL_EPOCH = date(1899, 12, 30)
MAX_AIRCREW_IMPORT_WARNINGS = 20
AIRCREW_TIME_PATTERN = re.compile(r"\d{1,2}:\d{2}\s*(?:[AaPp][Mm])?", re.IGNORECASE)
CUSTOM_SHIFT_TIME_PATTERN = re.compile(r"(\d{1,2})(?::(\d{2}))?\s*([AaPp][Mm])", re.IGNORECASE)
CREW_SHIFT_CUTOFF_MINUTES = (17 * 60) + 45


def _clean_header_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _normalize_carrier_label(value: Optional[str]) -> str:
    if not value:
        return ""
    label = str(value).strip()
    if not label:
        return ""
    compact = re.sub(r"[^A-Za-z]", "", label).upper()
    return AIRCREW_CARRIER_ALIAS_MAP.get(compact, label)


def _guess_carrier_name_from_sheet(sheet) -> Optional[str]:
    max_rows = min(sheet.max_row, 8)
    for row in sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        for cell in row:
            text = _clean_header_value(cell)
            if not text:
                continue
            low = text.lower()
            if "crew schedule" in low:
                pos = low.index("crew schedule")
                name = text[:pos].strip(" -–—:")
                normalized = _normalize_carrier_label(name)
                if normalized:
                    return normalized
            normalized = _normalize_carrier_label(text)
            if normalized in AIRCREW_CARRIERS:
                return normalized
    return None


def _coerce_excel_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        serial = int(value)
        if serial <= 0:
            return None
        try:
            return EXCEL_EPOCH + timedelta(days=serial)
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        text = re.sub(r"(\d)(st|nd|rd|th)", r"\1", text, flags=re.IGNORECASE)
        formats = [
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%m/%d/%y",
            "%B %d, %Y",
            "%b %d, %Y",
            "%B %d %Y",
            "%b %d %Y",
            "%B %d",
            "%b %d",
        ]
        for fmt in formats:
            try:
                parsed = datetime.strptime(text, fmt)
                if "%Y" not in fmt:
                    today = date.today()
                    parsed = parsed.replace(year=today.year)
                return parsed.date()
            except ValueError:
                continue
    return None


def _excel_fraction_to_time(value: float) -> Optional[str]:
    if value < 0:
        return None
    minutes_total = int(round(float(value) * 24 * 60))
    hour = (minutes_total // 60) % 24
    minute = minutes_total % 60
    return f"{hour:02d}:{minute:02d}"


def _extract_aircrew_times_from_cell(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, datetime):
        return [f"{value.hour:02d}:{value.minute:02d}"]
    if isinstance(value, timedelta):
        minutes_total = int(value.total_seconds() // 60)
        hour = (minutes_total // 60) % 24
        minute = minutes_total % 60
        return [f"{hour:02d}:{minute:02d}"]
    if isinstance(value, dt_time):
        return [f"{value.hour:02d}:{value.minute:02d}"]
    if isinstance(value, (int, float)):
        if 0 <= float(value) < 2:
            converted = _excel_fraction_to_time(float(value))
            return [converted] if converted else []
        text = str(value)
    else:
        text = str(value)
    text = text.strip()
    if not text:
        return []

    text = text.replace("\r", "\n")
    raw_tokens = re.split(r"[,\n;/]+", text)
    candidates: list[str] = []
    for token in raw_tokens:
        token = token.strip()
        if not token:
            continue
        if "-" in token:
            matches = AIRCREW_TIME_PATTERN.findall(token)
            if matches:
                candidates.extend(matches)
                continue
        candidates.append(token)
    if not candidates:
        candidates.extend(AIRCREW_TIME_PATTERN.findall(text))

    normalized: list[str] = []
    for token in candidates:
        cleaned = token.strip()
        if not cleaned:
            continue
        compact = cleaned.replace(" ", "").replace(".", "")
        no_minute = re.fullmatch(r"(\d{1,2})(AM|PM)", compact, flags=re.IGNORECASE)
        if no_minute:
            compact = f"{no_minute.group(1)}:00{no_minute.group(2)}"
        colonless = re.fullmatch(r"(\d{1,2})(\d{2})(AM|PM)?", compact, flags=re.IGNORECASE)
        if colonless and ":" not in compact:
            suffix = colonless.group(3) or ""
            compact = f"{colonless.group(1)}:{colonless.group(2)}{suffix}"
        normalized.append(compact)

    result: list[str] = []
    for token in normalized:
        try:
            result.append(_normalize_aircrew_time(token))
        except ValueError:
            continue
    return sorted(set(result))


def _parse_tabular_aircrew_sheet(workbook, forced_carrier: Optional[str] = None) -> tuple[dict[tuple[str, date], set[str]], list[str]]:
    sheet = workbook.active
    header_row_idx: Optional[int] = None
    headers: list[str] = []
    date_col_idx: Optional[int] = None
    max_scan = min(sheet.max_row, 25)
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        values = [_clean_header_value(cell) for cell in row]
        for col_idx, label in enumerate(values):
            if isinstance(label, str) and "date" in label.lower():
                header_row_idx = idx
                headers = values
                date_col_idx = col_idx
                break
        if header_row_idx is not None:
            break
    if header_row_idx is None or date_col_idx is None:
        raise ValueError("Could not locate a 'Date' column in the spreadsheet.")

    carrier_columns: list[tuple[int, str]] = []
    for idx, header in enumerate(headers):
        if idx == date_col_idx:
            continue
        label = header.strip()
        if not label:
            continue
        lowered = label.lower()
        if "date" in lowered or lowered in AIRCREW_HEADER_SKIP:
            continue
        carrier_columns.append((idx, label))

    if not carrier_columns:
        raise ValueError("No carrier columns were detected. Add at least one column with the carrier name.")
    if len(carrier_columns) == 1:
        guessed_name = forced_carrier or _guess_carrier_name_from_sheet(sheet)
        normalized = _normalize_carrier_label(guessed_name)
        if normalized:
            carrier_columns[0] = (carrier_columns[0][0], normalized)

    updates: dict[tuple[str, date], set[str]] = {}
    warnings: list[str] = []

    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        if not row or not any(cell not in (None, "") for cell in row):
            continue
        raw_date = row[date_col_idx] if date_col_idx < len(row) else None
        row_date = _coerce_excel_date(raw_date)
        if not row_date:
            if len(warnings) < MAX_AIRCREW_IMPORT_WARNINGS:
                warnings.append(f"Row {row_idx}: skipped because the Date cell is blank or invalid.")
            continue
        for col_idx, carrier in carrier_columns:
            value = row[col_idx] if col_idx < len(row) else None
            times = _extract_aircrew_times_from_cell(value)
            if not times:
                continue
            key = (carrier, row_date)
            updates.setdefault(key, set()).update(times)

    return updates, warnings


def _parse_vertical_aircrew_sheet(workbook, forced_carrier: Optional[str] = None) -> tuple[dict[tuple[str, date], set[str]], list[str]]:
    sheet = workbook.active
    header_row_idx: Optional[int] = None
    date_col_idx: Optional[int] = None
    eta_col_idx: Optional[int] = None
    max_scan = min(sheet.max_row, 25)
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        values = [_clean_header_value(cell) for cell in row]
        lowered = [val.lower() for val in values]
        for col_idx, label in enumerate(lowered):
            if date_col_idx is None and "date" in label:
                date_col_idx = col_idx
            if eta_col_idx is None and any(token in label for token in ("eta", "time", "arrival")):
                eta_col_idx = col_idx
        if date_col_idx is not None and eta_col_idx is not None:
            header_row_idx = idx
            break

    if header_row_idx is None or date_col_idx is None or eta_col_idx is None:
        raise ValueError("Could not locate both date and arrival time columns in this sheet.")

    carrier = forced_carrier or _guess_carrier_name_from_sheet(sheet)
    carrier = _normalize_carrier_label(carrier)
    if not carrier:
        raise ValueError("Could not determine which carrier this file belongs to. Choose one in the upload form.")

    updates: dict[tuple[str, date], set[str]] = {}
    warnings: list[str] = []
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        if not row or not any(cell not in (None, "") for cell in row):
            continue
        raw_date = row[date_col_idx] if date_col_idx < len(row) else None
        row_date = _coerce_excel_date(raw_date)
        if not row_date:
            if len(warnings) < MAX_AIRCREW_IMPORT_WARNINGS:
                warnings.append(f"Row {row_idx}: skipped because the Date cell is blank or invalid.")
            continue
        time_value = row[eta_col_idx] if eta_col_idx < len(row) else None
        times = _extract_aircrew_times_from_cell(time_value)
        if not times:
            continue
        key = (carrier, row_date)
        updates.setdefault(key, set()).update(times)
    if not updates:
        raise ValueError("No arrival times were found in the spreadsheet.")
    return updates, warnings


def _parse_aircrew_workbook(workbook, forced_carrier: Optional[str] = None) -> tuple[dict[tuple[str, date], set[str]], list[str]]:
    errors: list[str] = []
    try:
        return _parse_tabular_aircrew_sheet(workbook, forced_carrier)
    except ValueError as exc:
        errors.append(str(exc))
    try:
        return _parse_vertical_aircrew_sheet(workbook, forced_carrier)
    except ValueError as exc:
        errors.append(str(exc))
    raise ValueError(errors[-1] if errors else "Unable to parse the uploaded workbook.")


def _get_or_create_week(session: Session, start_d: date) -> Week:
    wk = session.scalar(select(Week).where(Week.start_date == start_d))
    if wk:
        return wk
    wk = Week(start_date=start_d)
    session.add(wk)
    session.flush()
    return wk


def _normalize_aircrew_time(value: str) -> str:
    token = (value or "").strip().upper()
    if not token:
        raise ValueError("Empty time")
    token = token.replace(" ", "")
    token = token.replace(".", "")
    token = token.replace("–", "-")
    match = re.fullmatch(r"(\d{1,2}):?(\d{2})(AM|PM)?", token)
    if not match:
        raise ValueError(f"Invalid time format: {value}")
    hour = int(match.group(1))
    minute = int(match.group(2))
    if hour > 23 or minute > 59:
        raise ValueError("Out of range time")
    suffix = match.group(3)
    if suffix:
        if hour > 12 or hour == 0:
            raise ValueError("Invalid 12-hour time")
        if hour == 12:
            hour = 0
        if suffix == "PM":
            hour += 12
    return f"{hour:02d}:{minute:02d}"


def _deserialize_aircrew_times(payload: Optional[str]) -> list[str]:
    if not payload:
        return []
    cleaned: list[str] = []
    def _append_normalized(value: str) -> None:
        try:
            cleaned.append(_normalize_aircrew_time(value))
        except Exception:
            pass

    try:
        data = json.loads(payload)
        if isinstance(data, list):
            for entry in data:
                if entry is None:
                    continue
                _append_normalized(str(entry))
            return sorted(set(cleaned))
    except Exception:
        pass

    # Legacy plain-text entries such as "10:15PM / 12:15AM"
    tokens = re.split(r"[,\n/]+", payload)
    for token in tokens:
        if token.strip():
            _append_normalized(token)
    return sorted(set(cleaned))


def _serialize_aircrew_times(times: Iterable[str]) -> str:
    unique = sorted({ _normalize_aircrew_time(t) for t in times if t is not None })
    return json.dumps(unique)


def _format_aircrew_time_display(value: str) -> str:
    try:
        hour, minute = map(int, value.split(":", 1))
    except Exception:
        return value
    suffix = "am" if hour < 12 else "pm"
    display_hour = hour % 12 or 12
    return f"{display_hour}:{minute:02d}{suffix}"


def _format_minutes_clock(total_minutes: int) -> str:
    total_minutes = total_minutes % (24 * 60)
    hour = total_minutes // 60
    minute = total_minutes % 60
    suffix = "am" if hour < 12 else "pm"
    display_hour = hour % 12 or 12
    return f"{display_hour}:{minute:02d}{suffix}"


def _format_aircrew_shift_window(start_minutes: int, end_minutes: int) -> str:
    start_label = _format_minutes_clock(start_minutes)
    end_label = _format_minutes_clock(end_minutes)
    return f"{start_label} - {end_label}"


def _suggest_shuttle_shift(minutes: list[int]) -> Optional[str]:
    if not minutes:
        return None
    ordered = sorted(minutes)
    extended = ordered + [ordered[0] + 24 * 60]
    max_gap = -1
    gap_idx = 0
    for idx in range(len(ordered)):
        gap = extended[idx + 1] - extended[idx]
        if gap > max_gap:
            max_gap = gap
            gap_idx = idx
    start_total = extended[gap_idx + 1]
    end_total = extended[gap_idx]
    if end_total < start_total:
        end_total += 24 * 60
    buffer = 60
    start_total = max(start_total - buffer, 0)
    end_total = min(end_total + buffer, start_total + (18 * 60))
    return _format_aircrew_shift_window(start_total, end_total)


def next_sort_order_for_section(session: Session, section_id: int) -> int:
    max_order = session.scalar(select(func.max(Employee.sort_order)).where(Employee.section_id == section_id))
    return (max_order or -1) + 1


def ensure_employee_sort_orders(session: Session, section_ids: Optional[Iterable[int]] = None) -> None:
    """Assign sequential sort orders for the provided sections."""
    if section_ids is None:
        query = select(Section)
    else:
        ids = list(section_ids)
        if not ids:
            return
        query = select(Section).where(Section.id.in_(ids))
    sections = session.scalars(query).all()
    for sec in sections:
        employees = list(
            session.scalars(
                select(Employee)
                .where(Employee.section_id == sec.id)
                .order_by(Employee.sort_order.is_(None), Employee.sort_order, Employee.id)
            )
        )
        for idx, emp in enumerate(employees):
            if emp.sort_order is None or emp.sort_order != idx:
                emp.sort_order = idx


def init_db_once():
    Base.metadata.create_all(engine)
    with SessionLocal() as s:
        # Ensure new columns exist for employees (SQLite light migration)
        with engine.connect() as conn:
            cols = [row[1] for row in conn.exec_driver_sql("PRAGMA table_info(employees)").fetchall()]
            if "availability" not in cols:
                conn.exec_driver_sql("ALTER TABLE employees ADD COLUMN availability TEXT")
            if "preferred_shift" not in cols:
                conn.exec_driver_sql("ALTER TABLE employees ADD COLUMN preferred_shift TEXT")
            if "seniority" not in cols:
                conn.exec_driver_sql("ALTER TABLE employees ADD COLUMN seniority INTEGER")
            to_cols = [row[1] for row in conn.exec_driver_sql("PRAGMA table_info(timeoff)").fetchall()]
            if "vacation" not in to_cols:
                conn.exec_driver_sql("ALTER TABLE timeoff ADD COLUMN vacation INTEGER DEFAULT 0")
            if "preferred_shifts_per_week" not in cols:
                conn.exec_driver_sql("ALTER TABLE employees ADD COLUMN preferred_shifts_per_week INTEGER")
            if "max_shifts_per_week" not in cols:
                conn.exec_driver_sql("ALTER TABLE employees ADD COLUMN max_shifts_per_week INTEGER")
            if "sort_order" not in cols:
                conn.exec_driver_sql("ALTER TABLE employees ADD COLUMN sort_order INTEGER")
            # Create aircrew arrivals table if missing
            conn.exec_driver_sql(
                """
                CREATE TABLE IF NOT EXISTS aircrew_arrivals (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    week_id INTEGER NOT NULL,
                    carrier TEXT NOT NULL,
                    date DATE NOT NULL,
                    times TEXT DEFAULT '',
                    UNIQUE(week_id, carrier, date)
                )
                """
            )
            # Create employee_roles table if missing
            conn.exec_driver_sql(
                """
                CREATE TABLE IF NOT EXISTS employee_roles (
                    id INTEGER PRIMARY KEY,
                    employee_id INTEGER NOT NULL,
                    section_id INTEGER NOT NULL,
                    FOREIGN KEY(employee_id) REFERENCES employees(id),
                    FOREIGN KEY(section_id) REFERENCES sections(id)
                )
                """
            )
            conn.exec_driver_sql(
                """
                CREATE TABLE IF NOT EXISTS occupancy_levels (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    week_id INTEGER NOT NULL,
                    date DATE NOT NULL,
                    percentage INTEGER,
                    UNIQUE(week_id, date)
                )
                """
            )
            conn.exec_driver_sql(
                """
                CREATE TABLE IF NOT EXISTS schedule_templates (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    slot INTEGER UNIQUE NOT NULL,
                    name TEXT NOT NULL,
                    payload TEXT NOT NULL,
                    saved_week_start DATE,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        # Ensure new column exists for assignments.dismissed_timeoff (SQLite light migration)
        with engine.connect() as conn:
            a_cols = [row[1] for row in conn.exec_driver_sql("PRAGMA table_info(assignments)").fetchall()]
            if "dismissed_timeoff" not in a_cols:
                conn.exec_driver_sql("ALTER TABLE assignments ADD COLUMN dismissed_timeoff INTEGER DEFAULT 0")

        # Ensure sections exist (and update FD required to 6)
        names = {
            "Breakfast Bar": None,
            "Front Desk": 6,
            "Shuttle": None,
            "Maintenance": 1,
        }
        existing = {sec.name: sec for sec in s.scalars(select(Section))}
        for n, required in names.items():
            if n in existing:
                if required is not None and (existing[n].required_per_day or 0) != required:
                    existing[n].required_per_day = required
            else:
                s.add(Section(name=n, required_per_day=required))
        s.flush()

        # Refresh sections after potential inserts
        sections = {sec.name: sec for sec in s.scalars(select(Section))}

        # Ensure baseline employees exist
        # def ensure_emp(name: str, sec_name: str):
        #     e = s.scalar(select(Employee).where(Employee.name == name))
        #     if not e:
        #         s.add(Employee(name=name, section_id=sections[sec_name].id))

        # for n in ["Rose", "Yoko", "Eurielle", "Anna", "Merve", "Ayako"]:
        #     ensure_emp(n, "Breakfast Bar")
        # for n in ["Ryan", "Emilyn", "Abdi", "Jordan", "Cindy"]:
        #     ensure_emp(n, "Front Desk")
        # for n in ["Alex", "Taylor", "Morgan"]:
        #     ensure_emp(n, "Shuttle")

        # # Seed time off if empty
        # if not s.scalar(select(TimeOff).limit(1)):
        #     to_items = [
        #         TimeOff(name="Christian", role="Front Desk", from_date=date(2025, 9, 10), to_date=date(2025, 9, 24), approved=True),
        #         TimeOff(name="Abdi", role="Front Desk", from_date=date(2025, 9, 18), to_date=date(2025, 9, 30), approved=True),
        #         TimeOff(name="Cindy", role="Front Desk", from_date=date(2025, 9, 18), to_date=date(2025, 9, 18), approved=False),
        #         TimeOff(name="Ayako", role="Breakfast Bar", from_date=date(2025, 9, 19), to_date=date(2025, 9, 19), approved=False),
        #         TimeOff(name="Merve", role="Breakfast Bar", from_date=date(2025, 9, 19), to_date=date(2025, 9, 19), approved=False),
        #         TimeOff(name="Tristan", role="Front Desk", from_date=date(2025, 9, 19), to_date=date(2025, 9, 20), approved=True),
        #     ]
        #     s.add_all(to_items)

        # Ensure there is a baseline week
        wk = s.scalar(select(Week).where(Week.start_date == date(2025, 9, 18)))
        if not wk:
            wk = Week(start_date=date(2025, 9, 18))
            s.add(wk)
            s.flush()

        # Ensure assignments exist for every employee/date
        day_list = daterange(wk.start_date, 7)
        for emp in s.scalars(select(Employee)):
            for d in day_list:
                exists = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == emp.id, Assignment.date == d))
                if not exists:
                    s.add(Assignment(week_id=wk.id, employee_id=emp.id, date=d, value="Set"))

        ensure_employee_sort_orders(s)

        s.commit()

        # Ensure TIME OFF is synchronized into assignments for approved requests
        sync_timeoff_to_assignments(wk.id, s)

        # Seed sample assignments if week is all Set
        any_non_set = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.value != "Set").limit(1))
        if not any_non_set:
            seed_example_assignments_db(wk.id, s)
            s.commit()


def has_approved_timeoff(name: str, role: str, dte: date, s: Session, exclude_id: Optional[int] = None) -> bool:
    stmt = (
        select(TimeOff)
        .where(
            TimeOff.name == name,
            TimeOff.role == role,
            TimeOff.approved.is_(True),
            TimeOff.from_date <= dte,
            TimeOff.to_date >= dte,
        )
    )
    if exclude_id is not None:
        stmt = stmt.where(TimeOff.id != exclude_id)
    to = s.scalar(stmt)
    return to is not None


def has_any_timeoff(name: str, role: str, dte: date, s: Session) -> bool:
    to = s.scalar(
        select(TimeOff).where(
            TimeOff.name == name,
            TimeOff.role == role,
            TimeOff.from_date <= dte,
            TimeOff.to_date >= dte,
        )
    )
    return to is not None


def _update_assignments_for_timeoff(
    s: Session,
    *,
    employee: Optional[Employee],
    start: date,
    end: date,
    approved: bool,
    timeoff: Optional[TimeOff] = None,
    exclude_id: Optional[int] = None,
) -> None:
    """Apply the time-off approval state to existing assignments across all weeks."""
    if not employee:
        return
    section = s.get(Section, employee.section_id)
    role_name = section.name if section else ""
    assignments = s.scalars(
        select(Assignment).where(
            Assignment.employee_id == employee.id,
            Assignment.date >= start,
            Assignment.date <= end,
        )
    )
    for assignment in assignments:
        if approved:
            desired = REQ_VAC_LABEL if timeoff and timeoff.vacation else TIME_OFF_LABEL
            if assignment.value != desired:
                assignment.value = desired
            if hasattr(Assignment, "dismissed_timeoff"):
                assignment.dismissed_timeoff = 0
        else:
            if assignment.value in TIME_OFF_VALUES and not has_approved_timeoff(employee.name, role_name, assignment.date, s, exclude_id):
                assignment.value = "Set"
                if hasattr(Assignment, "dismissed_timeoff"):
                    assignment.dismissed_timeoff = 0


def sync_timeoff_to_assignments(week_id: int, s: Session):
    wk = s.get(Week, week_id)
    days = list(daterange(wk.start_date, 7))
    # For each employee and day, if approved time off, set TIME OFF label
    for emp in s.scalars(select(Employee)):
        sec = s.get(Section, emp.section_id)
        role_name = sec.name if sec else ""
        for d in days:
            to_rec = s.scalar(
                select(TimeOff)
                .where(
                    TimeOff.name == emp.name,
                    TimeOff.role == role_name,
                    TimeOff.approved.is_(True),
                    TimeOff.from_date <= d,
                    TimeOff.to_date >= d,
                )
                .order_by(TimeOff.vacation.desc(), TimeOff.to_date.desc(), TimeOff.id.desc())
            )
            if to_rec:
                a = s.scalar(select(Assignment).where(Assignment.week_id == week_id, Assignment.employee_id == emp.id, Assignment.date == d))
                if a:
                    desired = REQ_VAC_LABEL if to_rec.vacation else TIME_OFF_LABEL
                    if a.value != desired:
                        a.value = desired
    s.commit()


def seed_example_assignments_db(week_id: int, s: Session):
    # Breakfast Bar, Front Desk, Shuttle examples
    bb_emp_ids = [e.id for e in s.scalars(select(Employee).join(Section).where(Section.name == "Breakfast Bar"))]
    fd_emp_ids = [e.id for e in s.scalars(select(Employee).join(Section).where(Section.name == "Front Desk"))]
    sh_emp_ids = [e.id for e in s.scalars(select(Employee).join(Section).where(Section.name == "Shuttle"))]
    maint_emp_ids = [e.id for e in s.scalars(select(Employee).join(Section).where(Section.name == "Maintenance"))]
    wk = s.get(Week, week_id)
    section_names = {sec.id: sec.name for sec in s.scalars(select(Section))}
    for d in daterange(wk.start_date, 7):
        # Breakfast Bar
        for eid in bb_emp_ids[:2]:  # only a couple for color variety
            emp = s.get(Employee, eid)
            role_name = section_names.get(emp.section_id, "") if emp else ""
            if emp and has_any_timeoff(emp.name, role_name, d, s):
                continue
            a = s.scalar(select(Assignment).where(Assignment.week_id == week_id, Assignment.employee_id == eid, Assignment.date == d))
        if a:
            a.value = choice(["5AM–12PM", "6AM–12PM", "7AM–12PM", "Set"])  # greens/blues
        # Front Desk: sample among AM/PM/Audit staggered options
        for eid in fd_emp_ids:
            emp = s.get(Employee, eid)
            role_name = section_names.get(emp.section_id, "") if emp else ""
            if emp and has_any_timeoff(emp.name, role_name, d, s):
                continue
            a = s.scalar(select(Assignment).where(Assignment.week_id == week_id, Assignment.employee_id == eid, Assignment.date == d))
            if a:
                a.value = choice([
                    "AM (6:00AM–2:00PM)", "AM (6:15AM–2:15PM)",
                    "PM (2:00PM–10:00PM)", "PM (2:15PM–10:15PM)",
                    "Audit (10:00PM–6:00AM)", "Audit (10:15PM–6:15AM)",
                    "Set"
                ])

        # Shuttle
        for eid in sh_emp_ids:
            emp = s.get(Employee, eid)
            role_name = section_names.get(emp.section_id, "") if emp else ""
            if emp and has_any_timeoff(emp.name, role_name, d, s):
                continue
            a = s.scalar(select(Assignment).where(Assignment.week_id == week_id, Assignment.employee_id == eid, Assignment.date == d))
            if a:
                shuttle_choices = [
                    "AM (3:30AM–11:30AM)",
                    "Midday (10:30AM–6:30PM)",
                    "PM (5:30PM–1:30AM)",
                    "Set",
                ] + SHUTTLE_CREW_SHIFTS
                a.value = choice(shuttle_choices)

        # Maintenance (one sample per day if available)
        if maint_emp_ids:
            pick = choice(maint_emp_ids)
            emp = s.get(Employee, pick)
            role_name = section_names.get(emp.section_id, "") if emp else ""
            if emp and not has_any_timeoff(emp.name, role_name, d, s):
                a = s.scalar(select(Assignment).where(Assignment.week_id == week_id, Assignment.employee_id == pick, Assignment.date == d))
                if a:
                    a.value = choice(["8AM–4:30PM", "Set"])


def has_generated_schedule(week_id: int) -> bool:
    """Check if the schedule has already been generated (has non-'Set' assignments)"""
    with SessionLocal() as s:
        any_non_set = s.scalar(select(Assignment).where(Assignment.week_id == week_id, Assignment.value != "Set").limit(1))
        return any_non_set is not None


def coverage_snapshot_db(week_id: int) -> tuple[dict, dict, int, dict, dict, int, dict, dict, int, dict, dict, dict, int, dict]:
    with SessionLocal() as s:
        wk = s.get(Week, week_id)
        dates = [d.isoformat() for d in daterange(wk.start_date, 7)]
        
        # Initialize Front Desk counts (2 per variant required)
        shift_variants = ["AM", "PM", "Audit"]
        counts = {k: {variant: 0 for variant in shift_variants} for k in dates}
        missing = {k: False for k in dates}

        # Initialize Shuttle counts (1 per variant required)
        sh_variants = ["AM", "Midday", "PM", "Crew"]
        sh_counts = {k: {variant: 0 for variant in sh_variants} for k in dates}
        sh_missing = {k: False for k in dates}

        # Initialize Breakfast counts (1 per variant required)
        bb_variants = ["5AM–12PM", "6AM–12PM", "7AM–12PM"]
        bb_counts = {k: {variant: 0 for variant in bb_variants} for k in dates}
        bb_missing = {k: False for k in dates}

        # Initialize Maintenance counts (1 per day required)
        maint_counts = {k: 0 for k in dates}
        maint_missing = {k: False for k in dates}

        # Track Front Desk duplicates of exact staggered times per variant (e.g., two at 2:15PM)
        fd_duplicates = {k: False for k in dates}
        # Exact label counts per date and variant for Front Desk
        fd_label_counts: dict[str, dict[str, dict[str, int]]] = {k: {"AM": {}, "PM": {}, "Audit": {}} for k in dates}
        
        # Count Front Desk assignments per shift variant per day
        # Include any employee assigned to a Front Desk-like label (AM/PM/Audit),
        # regardless of primary role, to account for secondary-role coverage.
        rows = s.scalars(select(Assignment).where(Assignment.week_id == week_id))

        for a in rows:
            if not a.value or a.value in NEUTRAL_ASSIGNMENT_VALUES:
                continue

            date_key = a.date.isoformat()

            # Front Desk variants: count ONLY if the value is a Front Desk label
            if a.value in FRONT_DESK_SHIFTS:
                if a.value.startswith("AM"):
                    counts[date_key]["AM"] += 1
                    fd_label_counts[date_key]["AM"][a.value] = fd_label_counts[date_key]["AM"].get(a.value, 0) + 1
                elif a.value.startswith("PM"):
                    counts[date_key]["PM"] += 1
                    fd_label_counts[date_key]["PM"][a.value] = fd_label_counts[date_key]["PM"].get(a.value, 0) + 1
                elif a.value.startswith("Audit"):
                    counts[date_key]["Audit"] += 1
                    fd_label_counts[date_key]["Audit"][a.value] = fd_label_counts[date_key]["Audit"].get(a.value, 0) + 1

            shuttle_value = a.value or ""
            # Shuttle variants: include dynamic crew suggestions so coverage stays accurate
            variant: Optional[str] = None
            if shuttle_value == SHUTTLE_COMBO_LABEL:
                sh_counts[date_key]["Midday"] += 1
                sh_counts[date_key]["Crew"] += 1
            else:
                variant = _infer_shuttle_variant(shuttle_value)
                if variant:
                    sh_counts[date_key][variant] += 1
                elif is_suggested_crew_label(shuttle_value):
                    sh_counts[date_key]["Crew"] += 1


            # Breakfast variants (exact labels)
            if a.value in bb_variants:
                bb_counts[date_key][a.value] += 1

            # Maintenance (single variant)
            if a.value == "8AM–4:30PM":
                maint_counts[date_key] += 1
        
        # Check for missing coverage: each variant needs at least 2 people
        for date_key in dates:
            for variant in shift_variants:
                if counts[date_key][variant] < 2:
                    missing[date_key] = True
                    break

        # Compute duplicate-stagger warnings for Front Desk per date
        for date_key in dates:
            dup_any = False
            for variant in ("AM", "PM", "Audit"):
                exact = fd_label_counts[date_key][variant]
                total = sum(exact.values())
                distinct = sum(1 for v in exact.values() if v > 0)
                if total >= 2 and distinct == 1:
                    dup_any = True
                    break
            fd_duplicates[date_key] = dup_any
        
        # For backward compatibility, also return total counts
        total_counts = {k: sum(counts[k].values()) for k in dates}
        required = 6  # 2 per variant * 3 variants
        # Shuttle missing and required (1 per each of 4 variants)
        for date_key in dates:
            for variant in sh_variants:
                if sh_counts[date_key][variant] < 1:
                    sh_missing[date_key] = True
                    break
        sh_required = 4

        # Breakfast missing and required (1 per each of 3 variants)
        for date_key in dates:
            for variant in bb_variants:
                if bb_counts[date_key][variant] < 1:
                    bb_missing[date_key] = True
                    break
        bb_required = 3

        # Maintenance missing threshold (1 per day)
        for date_key in dates:
            if maint_counts[date_key] < 1:
                maint_missing[date_key] = True
        maint_required = 1

        return (
            total_counts,
            missing,
            required,
            counts,
            sh_missing,
            sh_required,
            sh_counts,
            bb_missing,
            bb_required,
            bb_counts,
            fd_duplicates,
            maint_missing,
            maint_required,
            maint_counts,
        )


def double_booked_snapshot(week_id: int) -> dict[str, list[str]]:
    """Return date_key -> list of employee names double-booked (active in 2+ sections same day)."""
    with SessionLocal() as s:
        sections = list(s.scalars(select(Section)))
        sec_by_id = {sec.id: sec.name for sec in sections}
        emp_sections: dict[int, set[str]] = {}
        for e in s.scalars(select(Employee)):
            emp_sections.setdefault(e.id, set()).add(sec_by_id.get(e.section_id, ""))
        for er in s.scalars(select(EmployeeRole)):
            emp_sections.setdefault(er.employee_id, set()).add(sec_by_id.get(er.section_id, ""))

        multi_emp_ids = {eid for eid, secset in emp_sections.items() if len([n for n in secset if n]) > 1}
        if not multi_emp_ids:
            return {}

        def section_shifts(name: str) -> list[str]:
            if name == "Breakfast Bar":
                return BREAKFAST_SHIFTS
            if name == "Front Desk":
                return FRONT_DESK_SHIFTS
            if name == "Shuttle":
                return SHUTTLE_SHIFTS
            return []

        result: dict[str, list[str]] = {}
        wk = s.get(Week, week_id)
        if not wk:
            return {}
        for eid in multi_emp_ids:
            emp = s.get(Employee, eid)
            if not emp:
                continue
            for d in daterange(wk.start_date, 7):
                a = s.scalar(select(Assignment).where(Assignment.week_id == week_id, Assignment.employee_id == eid, Assignment.date == d))
                if not a or not a.value or a.value in NEUTRAL_ASSIGNMENT_VALUES:
                    continue
                active = 0
                val = a.value or ""
                for sec_name in (emp_sections.get(eid) or set()):
                    if sec_name and a.value in section_shifts(sec_name):
                        active += 1
                    elif sec_name == "Shuttle" and is_suggested_crew_label(val):
                        active += 1
                if active > 1:
                    key = d.isoformat()
                    result.setdefault(key, []).append(emp.name)
        return result


def build_week_context(week_id: int):
    with SessionLocal() as s:
        wk = s.get(Week, week_id)
        period_start, period_end = four_week_period_bounds(wk.start_date)
        week_start = wk.start_date
        week_end = week_start + timedelta(days=6)
        dates = week_dates(week_start)
        week_keys = {d["key"] for d in dates}
        # Sections and employees
        section_objs = list(s.scalars(select(Section)))
        sections = {
            sec.name: {
                "employees": [],
                "employee_ids": {},
                "assignments": {},
                "shifts": SECTION_SHIFT_MAP.get(sec.name, ["Set", TIME_OFF_LABEL, REQ_VAC_LABEL]),
                "employee_shifts": {},
            }
            for sec in section_objs
        }
        sec_by_id = {sec.id: sec.name for sec in section_objs}

        emp_primary_section: dict[int, str] = {}
        emp_sections_map: dict[int, set[str]] = {}

        for emp in s.scalars(select(Employee)):
            primary_name = sec_by_id.get(emp.section_id)
            if not primary_name:
                continue
            emp_primary_section[emp.id] = primary_name
            emp_sections_map.setdefault(emp.id, set()).add(primary_name)

        for er in s.scalars(select(EmployeeRole)):
            sec_name = sec_by_id.get(er.section_id)
            if not sec_name:
                continue
            emp_sections_map.setdefault(er.employee_id, set()).add(sec_name)

        def ordered_sections_for(emp_id: int) -> List[str]:
            available = emp_sections_map.get(emp_id) or set()
            if not available:
                return []
            primary = emp_primary_section.get(emp_id)
            ordered: List[str] = []
            if primary and primary in available:
                ordered.append(primary)
            for name in SECTION_DISPLAY_ORDER:
                if name in available and name != primary:
                    ordered.append(name)
            for name in available:
                if name not in ordered:
                    ordered.append(name)
            return ordered

        for sec_obj in section_objs:
            sec_name = sec_obj.name
            primary_emps = list(
                s.scalars(
                    select(Employee)
                    .where(Employee.section_id == sec_obj.id)
                    .order_by(Employee.sort_order.is_(None), Employee.sort_order, Employee.name)
                )
            )
            sections[sec_name]["employees"] = [e.name for e in primary_emps]

            for emp in primary_emps:
                sections[sec_name]["employee_ids"][emp.name] = emp.id
                sections[sec_name]["assignments"][emp.name] = {d["key"]: "Set" for d in dates}
                section_order = ordered_sections_for(emp.id) or [sec_name]
                sections[sec_name]["employee_shifts"][emp.name] = combined_shift_options(section_order)

            for emp in primary_emps:
                rows = s.scalars(select(Assignment).where(Assignment.week_id == week_id, Assignment.employee_id == emp.id))
                for a in rows:
                    sections[sec_name]["assignments"][emp.name][a.date.isoformat()] = a.value

        double_booked = double_booked_snapshot(wk.id)

        # Aircrew arrivals: carriers and per-day times arrays
        carriers_set = set(AIRCREW_CARRIERS)
        aircrew_rows = list(s.scalars(select(AircrewArrival).where(AircrewArrival.week_id == week_id)))
        for row in aircrew_rows:
            if row.carrier:
                carriers_set.add(row.carrier)
        carriers = sorted(carriers_set)
        arrivals: dict[str, dict[str, list[str]]] = {
            carrier: {d["key"]: [] for d in dates} for carrier in carriers
        }
        for row in aircrew_rows:
            carrier = row.carrier or ""
            key = row.date.isoformat()
            if carrier not in arrivals:
                arrivals[carrier] = {d["key"]: [] for d in dates}
                carriers.append(carrier)
            if key in arrivals[carrier]:
                arrivals[carrier][key] = _deserialize_aircrew_times(row.times)

        shuttle_suggestions: dict[str, str] = {}
        for d in dates:
            key = d["key"]
            all_minutes: list[int] = []
            for carrier_map in arrivals.values():
                for t in carrier_map.get(key, []):
                    try:
                        hour, minute = map(int, t.split(":", 1))
                    except Exception:
                        continue
                    all_minutes.append(hour * 60 + minute)
            suggestion = _suggest_shuttle_shift(all_minutes)
            if suggestion:
                shuttle_suggestions[key] = suggestion

        occupancy_rows = list(s.scalars(select(OccupancySnapshot).where(OccupancySnapshot.week_id == week_id)))
        occupancy_values = {d["key"]: None for d in dates}
        for row in occupancy_rows:
            key = row.date.isoformat()
            if key in occupancy_values:
                occupancy_values[key] = row.percentage

        # time off
        to_list = []
        vacation_days: dict[str, set[str]] = {}
        dismissed_days: dict[str, set[str]] = {}
        # Only surface time off that overlaps the displayed four-week period
        timeoff_query = (
            select(TimeOff)
            .where(TimeOff.to_date >= week_start, TimeOff.from_date <= week_end)
            .order_by(TimeOff.from_date, TimeOff.to_date, TimeOff.name)
        )
        for t in s.scalars(timeoff_query):
            same_day = t.from_date == t.to_date
            label = t.from_date.strftime("%b %d") if same_day else f"{t.from_date.strftime('%b %d')} to {t.to_date.strftime('%b %d')}"
            to_list.append({
                "id": t.id,
                "name": t.name,
                "role": t.role,
                "label": label,
                "approved": t.approved,
                "vacation": bool(getattr(t, 'vacation', False)),
            })
            # Build per-employee per-day vacation map for current week
            if bool(getattr(t, 'vacation', False)):
                cur = t.from_date
                while cur <= t.to_date:
                    k = cur.isoformat()
                    if k in week_keys:
                        vacation_days.setdefault(t.name, set()).add(k)
                    cur += timedelta(days=1)

        # Build dismissed_days map from assignments flag for this week (guard if column exists)
        if hasattr(Assignment, 'dismissed_timeoff'):
            for a in s.scalars(select(Assignment).where(Assignment.week_id == wk.id, Assignment.dismissed_timeoff == 1)):
                emp = s.get(Employee, a.employee_id)
                if not emp:
                    continue
                dismissed_days.setdefault(emp.name, set()).add(a.date.isoformat())

        # Check if schedule has been generated
        schedule_generated = has_generated_schedule(week_id)
        
        meta = {
            "range_label": format_four_week_label(period_start, period_end),
            "success_banner": (
                "4-week schedule saved from "
                f"{period_start.strftime('%B')} {period_start.day}, {period_start.year} "
                "to "
                f"{period_end.strftime('%B')} {period_end.day}, {period_end.year}!"
            ),
            "week_label": f"{format_week_label(wk.start_date)}",
            "fd_note": "Front Desk: 2 agents per AM/PM/Audit (6 total/day)",
            "schedule_generated": schedule_generated,
            "week_id": week_id,
        }

        carriers = sorted(arrivals.keys())
        template_slots = _template_slots_info(s)

        return {
            "week": {"label": "Week 1", "dates": dates, "sections": sections},
            "breakfast": sections["Breakfast Bar"],
            "front_desk": sections["Front Desk"],
            "meta": meta,
            "time_off": to_list,
            "vacation_days": vacation_days,
            "dismissed_days": dismissed_days,
            "double_booked": double_booked,
            "aircrew": {
                "carriers": carriers,
                "arrivals": arrivals,
            },
            "shuttle_suggestions": shuttle_suggestions,
            "occupancy": {
                "values": occupancy_values,
            },
            "schedule_templates": template_slots,
        }


def _with_template_upload_meta(meta: dict[str, Any]) -> dict[str, Any]:
    status = (request.args.get("template_status") or "").lower()
    if status not in {"success", "error"}:
        return meta
    message = (request.args.get("template_message") or "").strip()
    if not message:
        message = (
            "Schedule template updated successfully."
            if status == "success"
            else "Unable to update the schedule template."
        )
    enriched = dict(meta)
    enriched["template_upload"] = {"status": status, "message": message}
    return enriched


# Initialize DB at import time (Flask 3.x removed before_first_request)
init_db_once()

# ---- Routes ----


@app.route("/")
def index():
    today = date.today()
    with SessionLocal() as s:
        start = week_start_for_date(today)
        week = s.scalar(select(Week).where(Week.start_date == start))
        if not week:
            week = _ensure_week_and_assignments(s, start)
        ctx = build_week_context(week.id)
    meta = _with_template_upload_meta(ctx["meta"])
    (
        counts,
        missing,
        required,
        variant_counts,
        shuttle_missing,
        shuttle_required,
        shuttle_counts,
        bb_missing,
        bb_required,
        bb_counts,
        fd_duplicates,
        maintenance_missing,
        maintenance_required,
        maintenance_counts,
    ) = coverage_snapshot_db(week.id)
    return render_template(
        "schedule.html",
        meta=meta,
        week=ctx["week"],
        breakfast=ctx["breakfast"],
        front_desk=ctx["front_desk"],
        shuttle=ctx["week"]["sections"].get("Shuttle"),
        maintenance=ctx["week"]["sections"].get("Maintenance"),
        aircrew=ctx.get("aircrew", {}),
        occupancy=ctx.get("occupancy", {}),
        shuttle_suggestions=ctx.get("shuttle_suggestions", {}),
        counts=counts,
        missing=missing,
        required=required,
        variant_counts=variant_counts,
        shuttle_missing=shuttle_missing,
        shuttle_required=shuttle_required,
        shuttle_counts=shuttle_counts,
        bb_missing=bb_missing,
        bb_required=bb_required,
        bb_counts=bb_counts,
        maintenance_missing=maintenance_missing,
        maintenance_required=maintenance_required,
        maintenance_counts=maintenance_counts,
        fd_duplicates=fd_duplicates,
        time_off=ctx["time_off"],
        vacation_days=ctx.get("vacation_days", {}),
        dismissed_days=ctx.get("dismissed_days", {}),
        double_booked=ctx["double_booked"],
        schedule_templates=ctx.get("schedule_templates", []),
    )


@app.template_filter("fd_display")
def fd_display(label: str) -> str:
    # For Front Desk shift labels like "AM (6:00AM–2:00PM)", show only the time range
    if label in NEUTRAL_ASSIGNMENT_VALUES:
        return label
    if "(" in label and ")" in label:
        start = label.find("(") + 1
        end = label.find(")", start)
        if end > start:
            return label[start:end]
    return label


@app.template_filter("time_only")
def time_only(label: str) -> str:
    # For labels like "Midday (10:30AM–6:30PM)", show only the time range
    if label in NEUTRAL_ASSIGNMENT_VALUES:
        return label
    if "(" in label and ")" in label:
        start = label.find("(") + 1
        end = label.find(")", start)
        if end > start:
            return label[start:end]
    return label


def _shift_time_points(label: Optional[str]) -> list[int]:
    if not label:
        return []
    normalized = str(label).replace("–", "-")
    matches: list[int] = []
    for match in CUSTOM_SHIFT_TIME_PATTERN.finditer(normalized):
        hour = int(match.group(1))
        minute = int(match.group(2) or "0")
        period = (match.group(3) or "").lower()
        hour = hour % 12
        if period == "pm":
            hour += 12
        matches.append((hour * 60) + minute)
    return matches


def _shift_start_minutes(label: Optional[str]) -> Optional[int]:
    times = _shift_time_points(label)
    return times[0] if times else None


def _shift_window_minutes(label: Optional[str]) -> tuple[Optional[int], Optional[int]]:
    times = _shift_time_points(label)
    if len(times) < 2:
        return (None, None)
    start = times[0]
    end = times[1]
    if end <= start:
        end += 24 * 60
    return (start, end)


def _window_overlap_minutes(a_start: int, a_end: int, b_start: int, b_end: int) -> int:
    return max(0, min(a_end, b_end) - max(a_start, b_start))


@lru_cache(maxsize=1)
def _shift_classification_windows() -> list[tuple[int, int, str]]:
    windows: list[tuple[int, int, str]] = []
    for labels in SECTION_SHIFT_MAP.values():
        for label in labels:
            if not label or label in NEUTRAL_ASSIGNMENT_VALUES or label in TIME_OFF_VALUES or label == SHUTTLE_COMBO_LABEL:
                continue
            start, end = _shift_window_minutes(label)
            if start is None or end is None:
                continue
            css_class = _basic_shift_css_class(label)
            if not css_class:
                continue
            windows.append((start, end, css_class))
    return windows


def _match_shift_window_class(label: Optional[str]) -> str:
    if not label:
        return ""
    start, end = _shift_window_minutes(label)
    if start is None or end is None:
        return ""
    best_class = ""
    best_overlap = 0
    best_start = 1_000_000
    for win_start, win_end, css_class in _shift_classification_windows():
        overlap = _window_overlap_minutes(start, end, win_start, win_end)
        if overlap >= 300 and (
            overlap > best_overlap or (overlap == best_overlap and win_start < best_start)
        ):
            best_overlap = overlap
            best_start = win_start
            best_class = css_class
    return best_class


def _basic_shift_css_class(label: Optional[str]) -> str:
    if not label:
        return ""
    if label == "Set":
        return "select-gray"
    if label in TIME_OFF_VALUES:
        return "select-yellow"
    if label == "5AM–12PM":
        return "select-gold"
    if label == "6AM–12PM":
        return "select-blue"
    if label == "7AM–12PM":
        return "select-purple"
    if label == SHUTTLE_COMBO_LABEL:
        return "select-orange"
    if label.startswith("Audit"):
        return "select-red"
    if label.startswith("Crew"):
        return "select-red"
    if label.startswith("Midday"):
        return "select-blue"
    if label.startswith("PM (2"):
        return "select-blue"
    if label.startswith("PM"):
        return "select-purple"
    if label.startswith("AM"):
        return "select-green"
    if label == "8AM–4:30PM":
        return "select-green"
    start_minutes = _shift_start_minutes(label)
    if start_minutes is not None and start_minutes >= CREW_SHIFT_CUTOFF_MINUTES:
        return "select-red"
    return ""


def shift_css_class(label: Optional[str]) -> str:
    cls = _basic_shift_css_class(label)
    if cls:
        return cls
    return _match_shift_window_class(label)


@lru_cache(maxsize=1)
def _shuttle_variant_windows() -> dict[str, tuple[int, int]]:
    variants = {
        "AM": _shift_window_minutes("AM (3:30AM–11:30AM)"),
        "Midday": _shift_window_minutes("Midday (10:30AM–6:30PM)"),
        "PM": _shift_window_minutes("PM (5:30PM–1:30AM)"),
        "Crew": _shift_window_minutes(DEFAULT_CREW_SHIFT),
    }
    resolved: dict[str, tuple[int, int]] = {}
    for key, window in variants.items():
        if not window:
            continue
        start, end = window
        if start is None or end is None:
            continue
        resolved[key] = (start, end)
    return resolved


def _infer_shuttle_variant(label: Optional[str]) -> Optional[str]:
    value = (label or "").strip()
    if not value:
        return None
    if value == "AM (3:30AM–11:30AM)":
        return "AM"
    if value.startswith("Midday"):
        return "Midday"
    if value.startswith("PM (5:30PM"):
        return "PM"
    if value.startswith("Crew"):
        return "Crew"
    start_minutes = _shift_start_minutes(value)
    if start_minutes is not None and start_minutes >= CREW_SHIFT_CUTOFF_MINUTES:
        return "Crew"
    start, end = _shift_window_minutes(value)
    if start is None or end is None:
        return None
    best_variant = None
    best_overlap = 0
    for variant, (win_start, win_end) in _shuttle_variant_windows().items():
        overlap = _window_overlap_minutes(start, end, win_start, win_end)
        if overlap >= 300 and overlap > best_overlap:
            best_variant = variant
            best_overlap = overlap
    return best_variant


@app.template_filter("shift_class")
def shift_class_filter(label: str) -> str:
    return shift_css_class(label)


def combined_shift_options(section_names: Iterable[str]) -> List[str]:
    ordered_sections = list(section_names)
    if not ordered_sections:
        return ["Set", TIME_OFF_LABEL, REQ_VAC_LABEL]

    seen: set[str] = set()
    options: list[str] = []

    def _add(label: str) -> None:
        if label not in seen:
            options.append(label)
            seen.add(label)

    # Ensure neutral options appear first if any referenced section includes them
    for neutral in ("Set", TIME_OFF_LABEL, REQ_VAC_LABEL):
        if any(neutral in SECTION_SHIFT_MAP.get(sec, []) for sec in ordered_sections):
            _add(neutral)

    for sec in ordered_sections:
        for label in SECTION_SHIFT_MAP.get(sec, []):
            if label in ("Set", TIME_OFF_LABEL, REQ_VAC_LABEL):
                continue
            _add(label)

    return options or ["Set", TIME_OFF_LABEL, REQ_VAC_LABEL]


@app.template_filter("format_shift")
def format_shift(label: str) -> str:
    # Map Set -> "-" for compact display
    if not label:
        return label
    if label.lower() == "set":
        return "-"
    if label in TIME_OFF_VALUES:
        return label
    if label == SHUTTLE_COMBO_LABEL:
        return label
    if label.startswith("Crew Shift "):
        label = label[len("Crew Shift "):]
    # Prefer time-only inside parentheses if present
    if "(" in label and ")" in label:
        start = label.find("(") + 1
        end = label.find(")", start)
        if end > start:
            label = label[start:end]
    # Normalize dash spacing around en dash
    label = re.sub(r"\s*–\s*", " – ", label)
    # Also handle simple hyphen just in case
    label = re.sub(r"\s*-\s*", " - ", label)
    # Lowercase AM/PM tokens (handles attached forms like 6:00AM)
    label = re.sub(r"AM|PM", lambda m: m.group(0).lower(), label)
    return label


@app.template_filter("aircrew_time")
def aircrew_time_filter(value: Optional[str]) -> str:
    if not value:
        return ""
    return _format_aircrew_time_display(value)


@app.route("/week/<int:week_id>")
def view_week(week_id: int):
    ctx = build_week_context(week_id)
    meta = _with_template_upload_meta(ctx["meta"])
    (
        counts,
        missing,
        required,
        variant_counts,
        shuttle_missing,
        shuttle_required,
        shuttle_counts,
        bb_missing,
        bb_required,
        bb_counts,
        fd_duplicates,
        maintenance_missing,
        maintenance_required,
        maintenance_counts,
    ) = coverage_snapshot_db(week_id)
    return render_template(
        "schedule.html",
        meta=meta,
        week=ctx["week"],
        breakfast=ctx["breakfast"],
        front_desk=ctx["front_desk"],
        shuttle=ctx["week"]["sections"].get("Shuttle"),
        maintenance=ctx["week"]["sections"].get("Maintenance"),
        aircrew=ctx.get("aircrew", {}),
        occupancy=ctx.get("occupancy", {}),
        shuttle_suggestions=ctx.get("shuttle_suggestions", {}),
        counts=counts,
        missing=missing,
        required=required,
        variant_counts=variant_counts,
        shuttle_missing=shuttle_missing,
        shuttle_required=shuttle_required,
        shuttle_counts=shuttle_counts,
        bb_missing=bb_missing,
        bb_required=bb_required,
        bb_counts=bb_counts,
        maintenance_missing=maintenance_missing,
        maintenance_required=maintenance_required,
        maintenance_counts=maintenance_counts,
        fd_duplicates=fd_duplicates,
        time_off=ctx["time_off"],
        vacation_days=ctx.get("vacation_days", {}),
        dismissed_days=ctx.get("dismissed_days", {}),
        double_booked=ctx["double_booked"],
        schedule_templates=ctx.get("schedule_templates", []),
    )


def _fd_variant(label: str) -> Optional[str]:
    if not label:
        return None
    if label.startswith("AM"):
        return "AM"
    if label.startswith("PM"):
        return "PM"
    if label.startswith("Audit"):
        return "Audit"
    return None


@app.route("/week/<int:week_id>/manager-meals")
def manager_meals(week_id: int):
    # Build a map of day -> {variant -> list of employee names on that FD shift}
    with SessionLocal() as s:
        wk = s.get(Week, week_id)
        if not wk:
            return redirect(url_for("view_week", week_id=week_id))

        # Pull all assignments for the week and map by day/variant
        by_day: dict[date, dict[str, list[str]]] = {}
        rows = list(s.scalars(select(Assignment).where(Assignment.week_id == week_id)))
        # Preload employee id -> name for efficiency
        emp_name = {e.id: e.name for e in s.scalars(select(Employee))}
        for a in rows:
            if not a.value or a.value in NEUTRAL_ASSIGNMENT_VALUES:
                continue
            variant = _fd_variant(a.value)
            if not variant:
                continue
            # Only consider Front Desk-like labels
            if variant not in ("AM", "PM", "Audit"):
                continue
            nm = emp_name.get(a.employee_id)
            if not nm:
                continue
            by_day.setdefault(a.date, {}).setdefault(variant, []).append(nm)

        # Decide manager meal recipient per day/variant using seniority
        # AM Sunday-Thursday override goes to JoJo regardless of schedule
        results: list[dict] = []
        for d in daterange(wk.start_date, 7):
            entry = {
                "date": d,
                "label": d.strftime("%a %m/%d"),
                "AM": None,
                "PM": None,
                "Audit": None,
            }
            # AM override: Sunday (6) through Thursday (3)
            if d.weekday() in (6, 0, 1, 2, 3):
                entry["AM"] = "JoJo"
            else:
                am_list = by_day.get(d, {}).get("AM", [])
                entry["AM"] = next((n for n in SENIORITY_ORDER if n in am_list), (am_list[0] if am_list else None))
            # PM and Audit by seniority among assigned that shift
            for variant in ("PM", "Audit"):
                names = by_day.get(d, {}).get(variant, [])
                chosen = next((n for n in SENIORITY_ORDER if n in names), (names[0] if names else None))
                entry[variant] = chosen
            results.append(entry)

        # Build simple copy-friendly rows: Date -> "Night, AM, PM"
        rows: list[tuple[str, str]] = []
        for e in results:
            date_str = e["date"].strftime("%a %m/%d")
            audit = e.get("Audit") or "-"
            am = e.get("AM") or "-"
            pm = e.get("PM") or "-"
            rows.append((date_str, f"{audit}, {am}, {pm}"))

        return render_template(
            "meals.html",
            rows=rows,
            error=None,
        )


@app.route("/week/<int:week_id>/manager-meals.txt")
def manager_meals_text(week_id: int):
    # Plain-text export in format: Date: Audit, AM, PM
    with SessionLocal() as s:
        wk = s.get(Week, week_id)
        if not wk:
            return redirect(url_for("view_week", week_id=week_id))

        # Build assignments by day/variant
        by_day: dict[date, dict[str, list[str]]] = {}
        rows = list(s.scalars(select(Assignment).where(Assignment.week_id == week_id)))
        emp_name = {e.id: e.name for e in s.scalars(select(Employee))}
        for a in rows:
            if not a.value or a.value in NEUTRAL_ASSIGNMENT_VALUES:
                continue
            variant = _fd_variant(a.value)
            if not variant:
                continue
            if variant not in ("AM", "PM", "Audit"):
                continue
            nm = emp_name.get(a.employee_id)
            if not nm:
                continue
            by_day.setdefault(a.date, {}).setdefault(variant, []).append(nm)

        lines: list[str] = []
        for d in daterange(wk.start_date, 7):
            # Decide AM
            if d.weekday() in (6, 0, 1, 2, 3):
                am_name = "JoJo"
            else:
                am_list = by_day.get(d, {}).get("AM", [])
                am_name = next((n for n in SENIORITY_ORDER if n in am_list), (am_list[0] if am_list else None))
            # PM and Audit
            def pick(variant: str) -> Optional[str]:
                names = by_day.get(d, {}).get(variant, [])
                return next((n for n in SENIORITY_ORDER if n in names), (names[0] if names else None))
            audit_name = pick("Audit")
            pm_name = pick("PM")
            label = d.strftime("%a %m/%d")
            lines.append(f"{label}: {audit_name or '-'}, {am_name or '-'}, {pm_name or '-'}")

        text = "\n".join(lines)
        return Response(text, mimetype="text/plain")


    


@app.route("/admin/employees/add", methods=["GET", "POST"]) 
def admin_add_employee():
    if request.method == "GET":
        return render_template("admin_add_employee.html")
    # POST
    first_name = (request.form.get("first_name") or "").strip()
    last_name = (request.form.get("last_name") or "").strip()
    role = (request.form.get("role") or "").strip()
    if not first_name or not role:
        return render_template("admin_add_employee.html", error="First name and role are required"), 400
    with SessionLocal() as s:
        sec = s.scalar(select(Section).where(Section.name == role))
        if not sec:
            return render_template("admin_add_employee.html", error="Unknown role"), 400
        first_lower = first_name.casefold()
        last_lower = (last_name or "").casefold()
        same_first = s.scalar(
            select(func.count())
            .select_from(Employee)
            .where(
                Employee.section_id == sec.id,
                func.lower(Employee.first_name) == first_lower,
            )
        )
        if same_first and not last_name:
            return render_template(
                "admin_add_employee.html",
                error=f"Last name is required because {first_name} already exists in {role}.",
            ), 400
        duplicate = s.scalar(
            select(Employee)
            .where(
                Employee.section_id == sec.id,
                func.lower(Employee.first_name) == first_lower,
                func.lower(func.coalesce(Employee.last_name, "")) == last_lower,
            )
        )
        if duplicate:
            return render_template(
                "admin_add_employee.html",
                error=f"{first_name} {last_name or ''} is already in {role}.",
            ), 400
        full_name = format_employee_name(first_name, last_name or None)
        # Optional fields
        preferred_shift = (request.form.get("preferred_shift") or "").strip() or None
        seniority_raw = (request.form.get("seniority") or "").strip()
        seniority = int(seniority_raw) if seniority_raw.isdigit() else None
        pref_count_raw = (request.form.get("preferred_shifts_per_week") or "").strip()
        max_count_raw = (request.form.get("max_shifts_per_week") or "").strip()
        preferred_shifts_per_week = int(pref_count_raw) if pref_count_raw.isdigit() else None
        max_shifts_per_week = int(max_count_raw) if max_count_raw.isdigit() else None
        availability = (request.form.get("availability") or "").strip() or None
        sort_order = next_sort_order_for_section(s, sec.id)
        emp = Employee(
            name=full_name,
            section_id=sec.id,
            preferred_shift=preferred_shift,
            seniority=seniority,
            preferred_shifts_per_week=preferred_shifts_per_week,
            max_shifts_per_week=max_shifts_per_week,
            availability=availability,
            sort_order=sort_order,
            first_name=first_name,
            last_name=last_name or None,
        )
        s.add(emp)
        s.flush()
        # Ensure assignments for current week
        week = s.scalar(select(Week).where(Week.start_date == date(2025, 9, 18)))
        for d in daterange(week.start_date, 7):
            s.add(Assignment(week_id=week.id, employee_id=emp.id, date=d, value="Set"))
        s.commit()
    return redirect(url_for("list_employees"))


@app.route("/admin/employees")
def list_employees():
    with SessionLocal() as s:
        employees = {}
        sections = list(s.scalars(select(Section)))
        secondary_role_sets: dict[int, set[int]] = {}
        for eid, sid in s.execute(select(EmployeeRole.employee_id, EmployeeRole.section_id)):
            secondary_role_sets.setdefault(eid, set()).add(sid)
        secondary_role_map = {
            eid: sorted(list(sids))[0]
            for eid, sids in secondary_role_sets.items()
            if sids
        }
        role_options = [{"id": sec.id, "name": sec.name} for sec in sections]
        for sec in sections:
            employees[sec.name] = list(
                s.scalars(
                    select(Employee)
                    .where(Employee.section_id == sec.id)
                    .order_by(Employee.sort_order, Employee.name)
                )
            )
        section_ids = {sec.name: sec.id for sec in sections}
    shift_options = {
        "Breakfast Bar": [o for o in BREAKFAST_SHIFTS],
        "Front Desk": [o for o in FRONT_DESK_SHIFTS],
        "Shuttle": [o for o in SHUTTLE_SHIFTS],
        "Maintenance": [o for o in MAINTENANCE_SHIFTS],
    }
    return render_template(
        "employees.html",
        employees=employees,
        roles=sections,
        section_ids=section_ids,
        secondary_roles=secondary_role_map,
        role_options=role_options,
        shift_options=shift_options,
    )


@app.route("/admin/employees/reorder", methods=["POST"])
def reorder_employees():
    payload = request.get_json(silent=True) or {}
    section_id_raw = payload.get("section_id")
    employee_ids_raw = payload.get("employee_ids")
    try:
        section_id = int(section_id_raw)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Invalid section"}), 400
    if not isinstance(employee_ids_raw, list):
        return jsonify({"ok": False, "error": "Invalid employees"}), 400
    try:
        employee_ids = [int(eid) for eid in employee_ids_raw]
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Invalid employees"}), 400
    with SessionLocal() as s:
        section = s.get(Section, section_id)
        if not section:
            return jsonify({"ok": False, "error": "Section not found"}), 404
        db_ids = set(
            s.scalars(select(Employee.id).where(Employee.section_id == section_id))
        )
        if len(employee_ids) != len(db_ids) or set(employee_ids) != db_ids:
            return jsonify({"ok": False, "error": "Employees mismatch"}), 400
        for sort_index, eid in enumerate(employee_ids):
            emp = s.get(Employee, eid)
            if not emp or emp.section_id != section_id:
                continue
            emp.sort_order = sort_index
        ensure_employee_sort_orders(s, [section_id])
        s.commit()
    return jsonify({"ok": True})


def employee_roles_for(s: Session, eid: int) -> list[Section]:
    role_section_ids = [row[0] for row in s.execute(select(EmployeeRole.section_id).where(EmployeeRole.employee_id == eid)).all()]
    return list(s.scalars(select(Section).where(Section.id.in_(role_section_ids)))) if role_section_ids else []


@app.route("/admin/employees/<int:eid>/roles", methods=["GET", "POST"]) 
def manage_employee_roles(eid: int):
    with SessionLocal() as s:
        emp = s.get(Employee, eid)
        if not emp:
            return redirect(url_for('list_employees'))
        all_sections = list(s.scalars(select(Section)))
        primary_section_id = emp.section_id
        current_secondary = {er.section_id for er in s.scalars(select(EmployeeRole).where(EmployeeRole.employee_id == eid))}

        if request.method == "POST":
            if request.is_json:
                payload = request.get_json(silent=True) or {}
                raw_id = payload.get("secondary_role")
                try:
                    selected_id = int(raw_id)
                except (TypeError, ValueError):
                    if raw_id in (None, "", "none"):
                        selected_id = None
                    else:
                        return jsonify({"ok": False, "error": "Invalid role"}), 400
            else:
                raw_id = request.form.get("secondary_role")
                try:
                    selected_id = int(raw_id) if raw_id else None
                except (TypeError, ValueError):
                    selected_id = None
            if selected_id == primary_section_id:
                selected_id = None
            s.execute(delete(EmployeeRole).where(EmployeeRole.employee_id == eid))
            if selected_id:
                sec_exists = s.get(Section, selected_id)
                if not sec_exists:
                    if request.is_json:
                        return jsonify({"ok": False, "error": "Role not found"}), 404
                    selected_id = None
                else:
                    s.add(EmployeeRole(employee_id=eid, section_id=selected_id))
            s.commit()
            if request.is_json:
                return jsonify({"ok": True, "secondary_role": selected_id})
            return redirect(url_for('list_employees'))

        return render_template(
            "employee_roles.html",
            employee=emp,
            sections=all_sections,
            primary_id=primary_section_id,
            selected_ids=current_secondary,
        )


@app.route("/admin/employees/<int:eid>/role", methods=["POST"])
def change_employee_role(eid: int):
    payload = request.get_json(silent=True)
    is_json = payload is not None
    if is_json:
        section_id_raw = payload.get("section_id")
        try:
            section_id = int(section_id_raw)
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "Invalid role"}), 400
        role_name = None
    else:
        role_name = (request.form.get("role") or "").strip()
        if not role_name:
            return redirect(url_for('list_employees'))
        section_id = None
    with SessionLocal() as s:
        emp = s.get(Employee, eid)
        if not emp:
            if is_json:
                return jsonify({"ok": False, "error": "Employee not found"}), 404
            return redirect(url_for('list_employees'))
        if section_id is not None:
            sec = s.get(Section, section_id)
        else:
            sec = s.scalar(select(Section).where(Section.name == role_name))
        if not sec:
            if is_json:
                return jsonify({"ok": False, "error": "Role not found"}), 404
            return redirect(url_for('list_employees'))
        # Update role
        previous_section_id = emp.section_id
        emp.section_id = sec.id
        emp.sort_order = next_sort_order_for_section(s, sec.id)
        ensure_employee_sort_orders(s, [previous_section_id, sec.id])
        # Reset current week assignments to Set to avoid shift-type mismatch
        week = s.scalar(select(Week).where(Week.start_date == date(2025, 9, 18)))
        for d in daterange(week.start_date, 7):
            a = s.scalar(select(Assignment).where(Assignment.week_id == week.id, Assignment.employee_id == emp.id, Assignment.date == d))
            if a:
                a.value = "Set"
        s.commit()
        # Re-apply time off
        sync_timeoff_to_assignments(week.id, s)
    if is_json:
        return jsonify({"ok": True, "section_id": sec.id})
    return redirect(url_for('list_employees'))


@app.route("/admin/employees/<int:eid>/update", methods=["POST"])
def update_employee(eid: int):
    role = (request.form.get("role") or "").strip()
    availability = (request.form.get("availability") or "").strip()
    preferred_shift = (request.form.get("preferred_shift") or "").strip() or None
    seniority_raw = (request.form.get("seniority") or "").strip()
    seniority = int(seniority_raw) if seniority_raw.isdigit() else None
    with SessionLocal() as s:
        emp = s.get(Employee, eid)
        if not emp:
            return redirect(url_for('list_employees'))
        # Role change
        if role:
            sec = s.scalar(select(Section).where(Section.name == role))
            if sec and sec.id != emp.section_id:
                previous_section_id = emp.section_id
                emp.section_id = sec.id
                emp.sort_order = next_sort_order_for_section(s, sec.id)
                ensure_employee_sort_orders(s, [previous_section_id, sec.id])
                # Reset current week assignments to Set due to role change
                week = s.scalar(select(Week).where(Week.start_date == date(2025, 9, 18)))
                for d in daterange(week.start_date, 7):
                    a = s.scalar(select(Assignment).where(Assignment.week_id == week.id, Assignment.employee_id == emp.id, Assignment.date == d))
                    if a:
                        a.value = "Set"
        emp.availability = availability or None
        emp.preferred_shift = preferred_shift
        emp.seniority = seniority
        s.commit()
        # Re-apply time off after any changes
        week = s.scalar(select(Week).where(Week.start_date == date(2025, 9, 18)))
        sync_timeoff_to_assignments(week.id, s)
    return redirect(url_for('list_employees'))

@app.route("/admin/employees/<int:eid>/delete", methods=["POST"])
def delete_employee(eid: int):
    with SessionLocal() as s:
        emp = s.get(Employee, eid)
        if not emp:
            return redirect(url_for('list_employees'))
        section_id = emp.section_id
        # Delete assignments for this employee
        s.query(Assignment).filter(Assignment.employee_id == emp.id).delete()
        # Delete availability
        s.query(EmployeeAvailability).filter(EmployeeAvailability.employee_id == emp.id).delete()
        # Delete time off entries matching this name (name-based linkage)
        section = s.get(Section, emp.section_id)
        role_name = section.name if section else ""
        s.query(TimeOff).filter(TimeOff.name == emp.name, TimeOff.role == role_name).delete()
        # Delete employee
        s.delete(emp)
        ensure_employee_sort_orders(s, [section_id])
        s.commit()
    return redirect(url_for('list_employees'))


def role_shift_variants(role_name: str) -> list[str]:
    if role_name == "Breakfast Bar":
        return [s for s in BREAKFAST_SHIFTS if s not in NEUTRAL_ASSIGNMENT_VALUES]
    if role_name == "Front Desk":
        return [s for s in FRONT_DESK_SHIFTS if s not in NEUTRAL_ASSIGNMENT_VALUES]
    if role_name == "Shuttle":
        return [s for s in SHUTTLE_SHIFTS if s not in NEUTRAL_ASSIGNMENT_VALUES]
    if role_name == "Maintenance":
        return [s for s in MAINTENANCE_SHIFTS if s not in NEUTRAL_ASSIGNMENT_VALUES]
    return []


def role_availability_variants(role_name: str) -> list[str]:
    # Compact set for availability UI; Front Desk collapses to base variants
    if role_name == "Breakfast Bar":
        return ["5AM–12PM", "6AM–12PM", "7AM–12PM"]
    if role_name == "Front Desk":
        return ["AM", "PM", "Audit"]
    if role_name == "Shuttle":
        return [
            "AM (3:30AM–11:30AM)",
            "Midday (10:30AM–6:30PM)",
            "PM (5:30PM–1:30AM)",
        ] + SHUTTLE_CREW_SHIFTS + [SHUTTLE_COMBO_LABEL]
    if role_name == "Maintenance":
        return ["8AM–4:30PM"]
    return []


@app.route("/admin/employees/<int:eid>/availability", methods=["GET", "POST"]) 
def employee_availability(eid: int):
    days = [
        {"idx": 0, "label": "Mon"},
        {"idx": 1, "label": "Tue"},
        {"idx": 2, "label": "Wed"},
        {"idx": 3, "label": "Thu"},
        {"idx": 4, "label": "Fri"},
        {"idx": 5, "label": "Sat"},
        {"idx": 6, "label": "Sun"},
    ]
    with SessionLocal() as s:
        emp = s.get(Employee, eid)
        if not emp:
            return redirect(url_for('list_employees'))
        # Allow switching role context via query param for multi-role availability
        allowed_sections = [s.get(Section, emp.section_id)] + employee_roles_for(s, eid)
        allowed_names = [sec.name for sec in allowed_sections if sec]
        role = request.args.get("role") or (allowed_sections[0].name if allowed_sections else "")
        if role not in allowed_names and allowed_sections:
            role = allowed_sections[0].name
        variants = role_availability_variants(role)

        if request.method == "POST":
            # Save seniority and preferred shift
            seniority_raw = (request.form.get("seniority") or "").strip()
            emp.seniority = int(seniority_raw) if seniority_raw.isdigit() else None
            pref = (request.form.get("preferred_shift") or "").strip() or None
            # Only allow preferred shift that belongs to availability variants
            emp.preferred_shift = pref if (pref in variants) else None
            # Save weekly shift preferences
            pref_count_raw = (request.form.get("preferred_shifts_per_week") or "").strip()
            max_count_raw = (request.form.get("max_shifts_per_week") or "").strip()
            emp.preferred_shifts_per_week = int(pref_count_raw) if pref_count_raw.isdigit() else None
            emp.max_shifts_per_week = int(max_count_raw) if max_count_raw.isdigit() else None
            # Save availability: clear and re-add
            # NOTE: We don't clear all availability across roles; only rebuild for the role being edited.
            s.query(EmployeeAvailability).filter(
                EmployeeAvailability.employee_id == emp.id,
                EmployeeAvailability.shift_label.in_(variants)
            ).delete()
            selected = request.form.getlist("avail")  # values: "day::label"
            for token in selected:
                try:
                    d_str, label = token.split("::", 1)
                    d_idx = int(d_str)
                except ValueError:
                    continue
                if label in variants and 0 <= d_idx <= 6:
                    s.add(EmployeeAvailability(employee_id=emp.id, day_of_week=d_idx, shift_label=label, allowed=True))
            s.commit()
            return redirect(url_for('employee_availability', eid=eid, role=role))

        # GET: build checked set
        existing = {(ea.day_of_week, ea.shift_label) for ea in s.scalars(select(EmployeeAvailability).where(EmployeeAvailability.employee_id == emp.id))}
        # Render form
        return render_template(
            "employee_availability.html",
            employee=emp,
            role=role,
            variants=variants,
            days=days,
            existing=existing,
            allowed_roles=allowed_names,
        )


@app.route("/timeoff")
def timeoff_page():
    with SessionLocal() as s:
        items = []
        for t in s.scalars(select(TimeOff)):
            same_day = t.from_date == t.to_date
            label = t.from_date.strftime("%b %d") if same_day else f"{t.from_date.strftime('%b %d')} to {t.to_date.strftime('%b %d')}"
            items.append({"id": t.id, "name": t.name, "role": t.role, "label": label, "approved": t.approved, "vacation": bool(getattr(t, "vacation", False))})
        employees = list(s.scalars(select(Employee)))
        sections = list(s.scalars(select(Section)))
    return render_template("timeoff.html", time_off=items, employees=employees, sections=sections)


@app.route("/timeoff/new", methods=["GET", "POST"])
def timeoff_new():
    if request.method == 'GET':
        with SessionLocal() as s:
            employees = list(s.scalars(select(Employee)))
            sections = list(s.scalars(select(Section)))
        return render_template("timeoff_new.html", employees=employees, sections=sections)
    # POST
    name = (request.form.get('name') or '').strip()
    selected_role = (request.form.get('role') or '').strip()
    from_s = (request.form.get('from') or '').strip()
    to_s = (request.form.get('to') or '').strip()
    approved = bool(request.form.get('approved'))
    vacation = bool(request.form.get('vacation'))
    if not name or not selected_role or not from_s or not to_s:
        with SessionLocal() as s:
            employees = list(s.scalars(select(Employee)))
            sections = list(s.scalars(select(Section)))
        return render_template("timeoff_new.html", employees=employees, sections=sections, error="All fields are required"), 400
    try:
        fy, fm, fd = map(int, from_s.split('-'))
        ty, tm, td = map(int, to_s.split('-'))
        from_d, to_d = date(fy, fm, fd), date(ty, tm, td)
    except Exception:
        with SessionLocal() as s:
            employees = list(s.scalars(select(Employee)))
            sections = list(s.scalars(select(Section)))
        return render_template("timeoff_new.html", employees=employees, sections=sections, error="Invalid dates"), 400
    if to_d < from_d:
        with SessionLocal() as s:
            employees = list(s.scalars(select(Employee)))
            sections = list(s.scalars(select(Section)))
        return render_template("timeoff_new.html", employees=employees, sections=sections, error="End date must be after start date"), 400
    timeoff_info = None
    with SessionLocal() as s:
        emp = employee_by_role(s, name=name, role=selected_role)
        if not emp:
            employees = list(s.scalars(select(Employee)))
            sections = list(s.scalars(select(Section)))
            return render_template(
                "timeoff_new.html",
                employees=employees,
                sections=sections,
                error="No employee with that name for the selected role.",
            ), 400
        role_name = s.get(Section, emp.section_id).name if emp else 'Unknown'
        rec = TimeOff(name=name, role=role_name, from_date=from_d, to_date=to_d, approved=approved, vacation=vacation)
        s.add(rec)
        s.flush()
        timeoff_info = {
            "name": rec.name,
            "role": rec.role,
            "start": rec.from_date,
            "end": rec.to_date,
            "approved": bool(rec.approved),
            "vacation": bool(getattr(rec, "vacation", False)),
        }
        # If approved, update assignments for current week and broadcast
        if approved:
            _update_assignments_for_timeoff(
                s,
                employee=emp,
                start=from_d,
                end=to_d,
                approved=True,
                timeoff=rec,
            )
        s.commit()
        if approved:
            week = s.scalar(select(Week).where(Week.start_date == date(2025, 9, 18)))
            if week:
                (
                    counts,
                    missing,
                    required,
                    variant_counts,
                    shuttle_missing,
                    shuttle_required,
                    shuttle_counts,
                    bb_missing,
                    bb_required,
                    bb_counts,
                    fd_duplicates,
                    maintenance_missing,
                    maintenance_required,
                    maintenance_counts,
                ) = coverage_snapshot_db(week.id)
                _broadcast({
                    "type": "timeoff",
                    "op": "new",
                    "item": {
                        "id": rec.id,
                        "name": rec.name,
                        "role": rec.role,
                        "from": rec.from_date.isoformat(),
                        "to": rec.to_date.isoformat(),
                        "approved": True,
                        "vacation": bool(getattr(rec, "vacation", False)),
                    },
                    "counts": counts,
                    "missing": missing,
                    "required": required,
                    "variant_counts": variant_counts,
                    "shuttle_missing": shuttle_missing,
                    "shuttle_required": shuttle_required,
                    "shuttle_counts": shuttle_counts,
                    "bb_missing": bb_missing,
                    "bb_required": bb_required,
                    "bb_counts": bb_counts,
                    "fd_duplicates": fd_duplicates,
                    "maintenance_missing": maintenance_missing,
                    "maintenance_required": maintenance_required,
                    "maintenance_counts": maintenance_counts,
                    "double_booked": double_booked_snapshot(week.id),
                })
    if timeoff_info:
        _notify_timeoff_submission(**timeoff_info)
    return redirect(url_for('timeoff_page'))


@app.route("/schedules")
def schedules_page():
    with SessionLocal() as s:
        # Group weeks into 4-week periods anchored to the baseline Thursday
        periods: dict[date, dict] = {}
        for w in s.scalars(select(Week)):
            period_start, period_end = four_week_period_bounds(w.start_date)
            info = periods.setdefault(
                period_start,
                {
                    "first_week_id": w.id,
                    "start": period_start,
                    "end": period_end,
                },
            )
            # Ensure first_week_id points to the earliest week in the period
            existing_week = s.get(Week, info["first_week_id"]) if info.get("first_week_id") else None
            if existing_week and w.start_date < existing_week.start_date:
                info["first_week_id"] = w.id
        # Build list sorted by start date descending (most recent first)
        items = []
        for p in sorted(periods.values(), key=lambda x: x["start"], reverse=True):
            items.append({
                "id": p["first_week_id"],
                "start": p["start"].strftime('%b %d, %Y'),
                "end": p["end"].strftime('%b %d, %Y'),
            })
    _prune_expired_period_undos()
    undo_token = request.args.get("undo")
    undo_status = request.args.get("undo_status")
    undo_label = request.args.get("label")
    undo_context = None
    if undo_token:
        payload = _pending_period_undos.get(undo_token)
        if payload and payload.get("expires", 0) > time.time():
            seconds_left = max(0, int(payload["expires"] - time.time()))
            undo_context = {
                "token": undo_token,
                "label": payload.get("label"),
                "seconds": seconds_left,
                "expires": int(payload["expires"] * 1000),
            }
        else:
            undo_status = "expired"
            undo_label = None
    return render_template(
        "schedules.html",
        periods=items,
        undo_context=undo_context,
        undo_status=undo_status,
        undo_label=undo_label,
    )

@app.route("/assign", methods=["POST"])
def assign():
    data = request.get_json(force=True)
    section = data.get("section")
    employee_name = data.get("employee")
    date_key = data.get("date")
    value = data.get("value")
    week_id = data.get("week_id")  # Get week_id from request

    final_value = value or "Set"
    with SessionLocal() as s:
        # Use provided week_id or fall back to default week
        if week_id:
            week = s.get(Week, week_id)
        else:
            week = s.scalar(select(Week).where(Week.start_date == date(2025, 9, 18)))
        
        if not week:
            return jsonify({"ok": False, "error": "Week not found"}), 404
            
        # Resolve employee and allow secondary role assignments
        sec = s.scalar(select(Section).where(Section.name == section))
        if not sec:
            return jsonify({"ok": False, "error": "Unknown section"}), 400
        candidates = list(s.scalars(select(Employee).where(Employee.name == employee_name)))
        emp = next((cand for cand in candidates if cand.section_id == sec.id), None)
        if not emp:
            for cand in candidates:
                has_secondary = s.scalar(
                    select(EmployeeRole).where(
                        EmployeeRole.employee_id == cand.id,
                        EmployeeRole.section_id == sec.id,
                    ).limit(1)
                )
                if has_secondary:
                    emp = cand
                    break
        if not emp and candidates:
            emp = candidates[0]
        if not emp:
            return jsonify({"ok": False, "error": "Unknown employee"}), 400
        # Check permission: primary in section OR has secondary role mapping
        allowed = emp.section_id == sec.id or (s.scalar(select(EmployeeRole).where(EmployeeRole.employee_id == emp.id, EmployeeRole.section_id == sec.id).limit(1)) is not None)
        if not allowed:
            return jsonify({"ok": False, "error": "Employee not allowed for this section"}), 400
        try:
            y, m, d = map(int, date_key.split("-"))
            dte = date(y, m, d)
        except Exception:
            return jsonify({"ok": False, "error": "Bad date"}), 400

        a = s.scalar(select(Assignment).where(Assignment.week_id == week.id, Assignment.employee_id == emp.id, Assignment.date == dte))
        if not a:
            a = Assignment(week_id=week.id, employee_id=emp.id, date=dte, value="Set")
            s.add(a)
        # If there is approved time off on this date, allow manual override:
        # - When user selects a non-time-off value, record it and mark dismissed_timeoff=1
        #   to indicate a scheduled override despite approved time off.
        # - When user selects a time-off value (TIME OFF or REQ VAC), set it and clear the flag.
        if has_approved_timeoff(emp.name, sec.name, dte, s):
            if value and value not in TIME_OFF_VALUES:
                # Allow override with explicit shift
                a.value = value
                try:
                    # Best-effort; column may not exist in older DBs
                    a.dismissed_timeoff = True  # type: ignore[attr-defined]
                except Exception:
                    pass
            else:
                # Explicit time-off selection
                a.value = value if value in TIME_OFF_VALUES else TIME_OFF_LABEL
                try:
                    a.dismissed_timeoff = False  # type: ignore[attr-defined]
                except Exception:
                    pass
        else:
            # Normal assignment path; also clear dismissed flag if set previously
            a.value = value
            try:
                a.dismissed_timeoff = False  # type: ignore[attr-defined]
            except Exception:
                pass
        s.commit()
        final_value = a.value or "Set"

    (
        counts,
        missing,
        required,
        variant_counts,
        shuttle_missing,
        shuttle_required,
        shuttle_counts,
        bb_missing,
        bb_required,
        bb_counts,
        fd_duplicates,
        maintenance_missing,
        maintenance_required,
        maintenance_counts,
    ) = coverage_snapshot_db(week.id)
    response_payload = {
        "ok": True,
        "counts": counts,
        "missing": missing,
        "required": required,
        "variant_counts": variant_counts,
        "shuttle_missing": shuttle_missing,
        "shuttle_required": shuttle_required,
        "shuttle_counts": shuttle_counts,
        "bb_missing": bb_missing,
        "bb_required": bb_required,
        "bb_counts": bb_counts,
        "fd_duplicates": fd_duplicates,
        "maintenance_missing": maintenance_missing,
        "maintenance_required": maintenance_required,
        "maintenance_counts": maintenance_counts,
        "double_booked": double_booked_snapshot(week.id),
    }

    _notify_schedule_change(employee_name, section, dte, final_value)
    return jsonify(response_payload)


def _normalize_template_slot(raw_slot: Any) -> Optional[int]:
    try:
        slot = int(raw_slot)
    except (TypeError, ValueError):
        return None
    if slot < 1 or slot > TEMPLATE_SLOT_COUNT:
        return None
    return slot


@app.route("/schedule-templates/save", methods=["POST"])
def save_schedule_template():
    data = request.get_json(force=True) or {}
    slot = _normalize_template_slot(data.get("slot"))
    if slot is None:
        return jsonify({"ok": False, "error": "Invalid slot"}), 400
    try:
        week_id = int(data.get("week_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Invalid week"}), 400
    with SessionLocal() as s:
        week = s.get(Week, week_id)
        if not week:
            return jsonify({"ok": False, "error": "Week not found"}), 404
        payload = _capture_template_payload(s, week)
        template = s.scalar(select(ScheduleTemplate).where(ScheduleTemplate.slot == slot))
        if not template:
            template = ScheduleTemplate(
                slot=slot,
                name=_template_slot_label(slot),
                payload="{}",
                saved_week_start=week.start_date,
            )
            s.add(template)
        template.payload = json.dumps(payload)
        template.saved_week_start = week.start_date
        if not template.name:
            template.name = _template_slot_label(slot)
        template.updated_at = datetime.utcnow()
        s.commit()
        slot_info = _serialize_template_slot(slot, template)
    return jsonify({"ok": True, "slot": slot_info, "week_id": week_id})


@app.route("/schedule-templates/load", methods=["POST"])
def load_schedule_template():
    data = request.get_json(force=True) or {}
    slot = _normalize_template_slot(data.get("slot"))
    if slot is None:
        return jsonify({"ok": False, "error": "Invalid slot"}), 400
    try:
        week_id = int(data.get("week_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Invalid week"}), 400
    with SessionLocal() as s:
        week = s.get(Week, week_id)
        if not week:
            return jsonify({"ok": False, "error": "Week not found"}), 404
        template = s.scalar(select(ScheduleTemplate).where(ScheduleTemplate.slot == slot))
        if not template or not template.payload:
            return jsonify({"ok": False, "error": "This slot is empty"}), 404
        try:
            payload = json.loads(template.payload)
        except json.JSONDecodeError:
            return jsonify({"ok": False, "error": "Template data is corrupted"}), 500
        # Ensure assignments exist for this week before applying
        _ensure_week_and_assignments(s, week.start_date)
        applied = _apply_template_payload_to_week(payload, week, s)
        slot_info = _serialize_template_slot(slot, template)
    return jsonify({"ok": True, "slot": slot_info, "week_id": week_id, "applied": applied})


@app.route("/aircrew/arrival", methods=["POST"])
def upsert_aircrew_arrival():
    data = request.get_json(force=True)
    carrier = (data.get("carrier") or "").strip()
    date_key = (data.get("date") or "").strip()
    action = (data.get("action") or "add").strip().lower()
    time_value = data.get("time")
    week_id = data.get("week_id")
    try:
        week_id = int(week_id)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Invalid week id"}), 400
    if not carrier:
        return jsonify({"ok": False, "error": "Carrier is required"}), 400
    if not date_key:
        return jsonify({"ok": False, "error": "Date is required"}), 400
    try:
        y, m, d = map(int, date_key.split("-"))
        dte = date(y, m, d)
    except Exception:
        return jsonify({"ok": False, "error": "Bad date"}), 400

    with SessionLocal() as s:
        week = s.get(Week, week_id)
        if not week:
            return jsonify({"ok": False, "error": "Week not found"}), 404
        week_end = week.start_date + timedelta(days=6)
        if dte < week.start_date or dte > week_end:
            return jsonify({"ok": False, "error": "Date outside of selected week"}), 400
        response_cells: dict[str, list[str]] = {}

        def load_row(target_date: date) -> tuple[Optional[AircrewArrival], list[str]]:
            row = s.scalar(
                select(AircrewArrival).where(
                    AircrewArrival.week_id == week.id,
                    AircrewArrival.carrier == carrier,
                    AircrewArrival.date == target_date,
                )
            )
            return row, _deserialize_aircrew_times(row.times if row else "")

        def save_row(target_date: date, times_list: list[str], row: Optional[AircrewArrival]) -> None:
            normalized = sorted({t for t in times_list})
            if row and not normalized:
                s.delete(row)
                return
            if not row and not normalized:
                return
            if not row:
                row = AircrewArrival(week_id=week.id, carrier=carrier, date=target_date, times="[]")
                s.add(row)
            row.times = _serialize_aircrew_times(normalized)

        row, current_times = load_row(dte)
        if action == "remove":
            try:
                to_remove = _normalize_aircrew_time(time_value or "")
            except Exception:
                return jsonify({"ok": False, "error": "Invalid time"}), 400
            updated = [t for t in current_times if t != to_remove]
            save_row(dte, updated, row)
            response_cells[dte.isoformat()] = updated
        else:
            try:
                to_add = _normalize_aircrew_time(time_value or "")
            except Exception:
                return jsonify({"ok": False, "error": "Invalid time"}), 400
            if to_add not in current_times:
                current_times.append(to_add)
            save_row(dte, current_times, row)
            response_cells[dte.isoformat()] = sorted(set(current_times))

        s.commit()

    batch = [{"carrier": carrier, "date": dk, "times": times, "week_id": week_id} for dk, times in response_cells.items()]
    _broadcast({"type": "aircrew", "week_id": week_id, "batch": batch})
    return jsonify({
        "ok": True,
        "carrier": carrier,
        "week_id": week_id,
        "cells": response_cells,
    })


@app.route("/aircrew/import", methods=["POST"])
def import_aircrew_schedule():
    upload = request.files.get("file")
    if upload is None or not upload.filename:
        return jsonify({"ok": False, "error": "Select an Excel file to upload."}), 400
    preferred_carrier_raw = (request.form.get("carrier") or "").strip()
    preferred_carrier = _normalize_carrier_label(preferred_carrier_raw) if preferred_carrier_raw else None
    payload = upload.read()
    if not payload:
        return jsonify({"ok": False, "error": "The uploaded file is empty."}), 400
    try:
        workbook = load_workbook(io.BytesIO(payload), data_only=True)
    except Exception:
        return jsonify({"ok": False, "error": "Unable to read that Excel file. Please upload a .xlsx file."}), 400
    try:
        parsed_updates, warnings = _parse_aircrew_workbook(workbook, preferred_carrier)
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    if not parsed_updates:
        return jsonify({"ok": False, "error": "No arrival times were found in the spreadsheet."}), 400

    display_week_id = request.form.get("week_id")
    try:
        display_week_id_int = int(display_week_id)
    except (TypeError, ValueError):
        display_week_id_int = None

    touched_dates: set[date] = set()
    touched_carriers: set[str] = set()
    per_week_batches: dict[int, list[dict]] = defaultdict(list)
    week_cells: dict[str, dict[str, list[str]]] = {}
    updated_cells = 0

    with SessionLocal() as s:
        display_week = s.get(Week, display_week_id_int) if display_week_id_int else None
        for (carrier, day), times_set in parsed_updates.items():
            normalized_times = sorted(times_set)
            touched_dates.add(day)
            touched_carriers.add(carrier)
            week_start = week_start_for_date(day)
            week = _get_or_create_week(s, week_start)
            row = s.scalar(
                select(AircrewArrival).where(
                    AircrewArrival.week_id == week.id,
                    AircrewArrival.carrier == carrier,
                    AircrewArrival.date == day,
                )
            )
            existing_times = _deserialize_aircrew_times(row.times) if row else []
            if existing_times == normalized_times:
                continue
            serialized = _serialize_aircrew_times(normalized_times)
            if row:
                row.times = serialized
            else:
                row = AircrewArrival(week_id=week.id, carrier=carrier, date=day, times=serialized)
                s.add(row)
            updated_cells += 1
            entry = {
                "carrier": carrier,
                "date": day.isoformat(),
                "times": normalized_times,
            }
            per_week_batches[week.id].append(entry)
            if display_week and week.id == display_week.id:
                week_cells.setdefault(carrier, {})[day.isoformat()] = normalized_times
        s.commit()

    for wk_id, batch in per_week_batches.items():
        if batch:
            broadcast_batch = [{**item, "week_id": wk_id} for item in batch]
            _broadcast({"type": "aircrew", "week_id": wk_id, "batch": broadcast_batch})

    warnings = warnings[:MAX_AIRCREW_IMPORT_WARNINGS]
    response = {
        "ok": True,
        "updated_cells": updated_cells,
        "touched_dates": sorted(d.isoformat() for d in touched_dates),
        "touched_carriers": sorted(touched_carriers),
        "week_cells": week_cells,
        "warnings": warnings,
        "weeks_affected": sorted(per_week_batches.keys()),
    }
    if updated_cells == 0:
        response["message"] = "No changes were applied because the uploaded times match what is already saved."
    else:
        response["message"] = (
            f"Updated {updated_cells} arrival cell{'s' if updated_cells != 1 else ''} "
            f"across {len(touched_dates)} date{'s' if len(touched_dates) != 1 else ''}."
        )
    return jsonify(response)


@app.route("/aircrew/template.xlsx")
def aircrew_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Aircrew Import"
    headers = ["Date"] + list(AIRCREW_CARRIERS)
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)
        width = 16 if idx == 1 else 22
        ws.column_dimensions[get_column_letter(idx)].width = width
    today = date.today()
    start = today.replace(day=1)
    carrier_count = len(AIRCREW_CARRIERS)
    sample_rows = [
        ["10:15pm\n11:45pm", "8:30am"],
        ["2:05am", "6:15pm"],
        ["", "7:45am"],
        ["11:05pm", "11:15pm\n11:45pm"],
    ]
    for offset, row_samples in enumerate(sample_rows):
        dte = start + timedelta(days=offset)
        ws.cell(row=offset + 2, column=1, value=dte)
        ws.cell(row=offset + 2, column=1).number_format = "yyyy-mm-dd"
        row_values = list(row_samples) + [""] * max(0, carrier_count - len(row_samples))
        for col_offset, sample in enumerate(row_values[:carrier_count], start=2):
            ws.cell(row=offset + 2, column=col_offset, value=sample)
    ws.cell(row=len(sample_rows) + 3, column=1, value="Use commas or line breaks to list multiple arrival times in a day.")
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="aircrew-import-template.xlsx",
    )


def _parse_occupancy_report(payload: str) -> dict[date, int]:
    """Extract date -> occupancy percentage mapping from a text report."""
    results: dict[date, int] = {}
    for raw_line in payload.splitlines():
        line = (raw_line or "").strip()
        if not line:
            continue
        date_match = OCCUPANCY_DATE_RE.search(line)
        if not date_match:
            continue
        month, day, year = map(int, date_match.groups())
        year += 2000 if year < 70 else 1900
        try:
            dte = date(year, month, day)
        except ValueError:
            continue
        percent_match = OCCUPANCY_PERCENT_RE.search(line, date_match.end())
        if not percent_match:
            continue
        try:
            pct = float(percent_match.group(1))
        except ValueError:
            continue
        pct_int = max(0, min(100, int(round(pct))))
        results[dte] = pct_int
    return results


@app.route("/occupancy", methods=["POST"])
def update_occupancy():
    data = request.get_json(force=True)
    date_key = (data.get("date") or "").strip()
    week_id_raw = data.get("week_id")
    value_raw = data.get("value")

    try:
        week_id = int(week_id_raw)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Invalid week id"}), 400
    if not date_key:
        return jsonify({"ok": False, "error": "Date is required"}), 400
    try:
        y, m, d = map(int, date_key.split("-"))
        dte = date(y, m, d)
    except Exception:
        return jsonify({"ok": False, "error": "Bad date"}), 400

    def normalize_percentage(raw) -> Optional[int]:
        if raw is None:
            return None
        if isinstance(raw, str):
            raw = raw.strip()
            if not raw:
                return None
        try:
            numeric = float(raw)
        except (TypeError, ValueError):
            raise ValueError
        if math.isnan(numeric):
            raise ValueError
        numeric = max(0.0, min(100.0, numeric))
        return int(round(numeric))

    try:
        pct_value = normalize_percentage(value_raw)
    except ValueError:
        return jsonify({"ok": False, "error": "Percentage must be between 0 and 100"}), 400

    with SessionLocal() as s:
        week = s.get(Week, week_id)
        if not week:
            return jsonify({"ok": False, "error": "Week not found"}), 404
        week_end = week.start_date + timedelta(days=6)
        if dte < week.start_date or dte > week_end:
            return jsonify({"ok": False, "error": "Date outside of selected week"}), 400
        row = s.scalar(
            select(OccupancySnapshot).where(
                OccupancySnapshot.week_id == week.id,
                OccupancySnapshot.date == dte,
            )
        )
        if pct_value is None:
            if row:
                s.delete(row)
        else:
            if not row:
                row = OccupancySnapshot(week_id=week.id, date=dte, percentage=pct_value)
                s.add(row)
            else:
                row.percentage = pct_value
        s.commit()

    _broadcast({
        "type": "occupancy",
        "week_id": week_id,
        "items": [
            {
                "date": dte.isoformat(),
                "value": pct_value,
            }
        ],
    })
    return jsonify({"ok": True, "date": dte.isoformat(), "value": pct_value})


@app.route("/occupancy/import", methods=["POST"])
def import_occupancy_report():
    upload = request.files.get("file")
    if upload is None or not upload.filename:
        return jsonify({"ok": False, "error": "Select a report file to upload."}), 400
    payload = upload.read()
    if not payload:
        return jsonify({"ok": False, "error": "The uploaded file is empty."}), 400
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = payload.decode("latin-1", errors="ignore")
    parsed = _parse_occupancy_report(text)
    if not parsed:
        return jsonify({"ok": False, "error": "No dates with occupancy percentages were found in that file."}), 400

    display_week_id = request.form.get("week_id")
    try:
        display_week_id_int = int(display_week_id)
    except (TypeError, ValueError):
        display_week_id_int = None

    updated = 0
    touched_dates: set[str] = set()
    per_week_batches: dict[int, list[dict[str, Any]]] = defaultdict(list)
    week_cells: dict[str, Optional[int]] = {}

    with SessionLocal() as s:
        display_week = s.get(Week, display_week_id_int) if display_week_id_int else None
        for dte, pct_value in sorted(parsed.items()):
            week_start = week_start_for_date(dte)
            week = _get_or_create_week(s, week_start)
            row = s.scalar(
                select(OccupancySnapshot).where(
                    OccupancySnapshot.week_id == week.id,
                    OccupancySnapshot.date == dte,
                )
            )
            if row and row.percentage == pct_value:
                continue
            if row:
                row.percentage = pct_value
            else:
                row = OccupancySnapshot(week_id=week.id, date=dte, percentage=pct_value)
                s.add(row)
            updated += 1
            iso_key = dte.isoformat()
            touched_dates.add(iso_key)
            per_week_batches[week.id].append({"date": iso_key, "value": pct_value})
            if display_week and display_week.id == week.id:
                week_cells[iso_key] = pct_value
        s.commit()

    for wk_id, items in per_week_batches.items():
        if not items:
            continue
        _broadcast({"type": "occupancy", "week_id": wk_id, "items": items})

    if updated == 0:
        return jsonify({
            "ok": True,
            "updated_cells": 0,
            "touched_dates": sorted(touched_dates),
            "week_cells": week_cells,
            "message": "No changes were applied.",
        })

    return jsonify({
        "ok": True,
        "updated_cells": updated,
        "touched_dates": sorted(touched_dates),
        "week_cells": week_cells,
        "message": f"Updated {updated} day{'s' if updated != 1 else ''}.",
    })


def _guess_image_mime(upload) -> Optional[str]:
    mimetype = (getattr(upload, "mimetype", None) or "").strip()
    if mimetype and mimetype.startswith("image/"):
        return mimetype
    filename = getattr(upload, "filename", "") or ""
    guessed, _ = mimetypes.guess_type(filename)
    if guessed and guessed.startswith("image/"):
        return guessed
    return None


@app.route("/api/whatsapp/send-image", methods=["POST"])
def send_whatsapp_image():
    config = _twilio_config()
    if not config:
        return jsonify({
            "ok": False,
            "error": "WhatsApp sandbox is not configured. Set TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN, TWILIO_WHATSAPP_TO, and WHATSAPP_MEDIA_BASE_URL.",
        }), 503

    upload = request.files.get("image")
    if upload is None or not upload.filename:
        return jsonify({"ok": False, "error": "Paste an image before sending."}), 400

    file_bytes = upload.read() or b""
    if not file_bytes:
        return jsonify({"ok": False, "error": "Could not read the pasted image."}), 400
    if len(file_bytes) > WHATSAPP_MAX_IMAGE_BYTES:
        limit_mb = WHATSAPP_MAX_IMAGE_BYTES // (1024 * 1024)
        return jsonify({"ok": False, "error": f"Images must be smaller than {limit_mb} MB."}), 400

    mime_type = _guess_image_mime(upload)
    if not mime_type:
        mime_type = "application/octet-stream"
    if not mime_type.startswith("image/"):
        return jsonify({"ok": False, "error": "Only image files can be sent to WhatsApp."}), 400

    caption = (request.form.get("caption") or "").strip()
    guessed_ext = mimetypes.guess_extension(mime_type.split(";")[0]) or ".bin"
    stored_filename = _persist_whatsapp_media(file_bytes, guessed_ext)
    media_url = _build_media_url(config["media_base_url"], stored_filename)

    try:
        _send_twilio_whatsapp_image(config, media_url=media_url, caption=caption or None)
    except TwilioWhatsAppError as exc:
        app.logger.warning("Twilio WhatsApp error: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 502

    return jsonify({"ok": True, "message": "Image sent to WhatsApp sandbox."})


@app.route("/timeoff/delete/<int:tid>", methods=["POST"])
def delete_timeoff(tid):
    with SessionLocal() as s:
        to = s.get(TimeOff, tid)
        if not to:
            return jsonify({"ok": False, "error": "Not found"}), 404
        
        # Get the week reference for coverage updates
        week = s.scalar(select(Week).where(Week.start_date == date(2025, 9, 18)))
        
        # If it was approved, we need to update assignments back to "Set"
        if to.approved:
            emp = employee_by_role(s, name=to.name, role=to.role or "")
            _update_assignments_for_timeoff(
                s,
                employee=emp,
                start=to.from_date,
                end=to.to_date,
                approved=False,
                timeoff=to,
                exclude_id=to.id,
            )

        s.delete(to)
        s.commit()

        # Only update coverage if we have a valid week
        if week:
            (
                counts,
                missing,
                required,
                variant_counts,
                shuttle_missing,
                shuttle_required,
                shuttle_counts,
                bb_missing,
                bb_required,
                bb_counts,
                fd_duplicates,
                maintenance_missing,
                maintenance_required,
                maintenance_counts,
            ) = coverage_snapshot_db(week.id)
            # Broadcast deletion to live listeners
            _broadcast({
                "type": "timeoff",
                "op": "delete",
                "item": {
                    "id": tid,
                    "name": to.name,
                    "role": to.role,
                    "from": to.from_date.isoformat(),
                    "to": to.to_date.isoformat(),
                    "approved": bool(to.approved),
                    "vacation": bool(getattr(to, "vacation", False)),
                },
                "counts": counts,
                "missing": missing,
                "required": required,
                "variant_counts": variant_counts,
                "shuttle_missing": shuttle_missing,
                "shuttle_required": shuttle_required,
                "shuttle_counts": shuttle_counts,
                "bb_missing": bb_missing,
                "bb_required": bb_required,
                "bb_counts": bb_counts,
                "fd_duplicates": fd_duplicates,
                "maintenance_missing": maintenance_missing,
                "maintenance_required": maintenance_required,
                "maintenance_counts": maintenance_counts,
                "double_booked": double_booked_snapshot(week.id),
            })
            return jsonify({
                "ok": True,
                "counts": counts,
                "missing": missing,
                "required": required,
                "variant_counts": variant_counts,
                "shuttle_missing": shuttle_missing,
                "shuttle_required": shuttle_required,
                "shuttle_counts": shuttle_counts,
                "bb_missing": bb_missing,
                "bb_required": bb_required,
                "bb_counts": bb_counts,
                "fd_duplicates": fd_duplicates,
                "maintenance_missing": maintenance_missing,
                "maintenance_required": maintenance_required,
                "maintenance_counts": maintenance_counts,
                "double_booked": double_booked_snapshot(week.id),
            })
        else:
            return jsonify({"ok": True, "counts": {}, "missing": {}, "required": 0, "variant_counts": {}})


@app.route("/timeoff/toggle", methods=["POST"])
def toggle_timeoff():
    data = request.get_json(force=True)
    tid = data.get("id")
    approved = bool(data.get("approved"))
    with SessionLocal() as s:
        to = s.get(TimeOff, tid)
        if not to:
            return jsonify({"ok": False, "error": "Not found"}), 404
        to.approved = approved
        emp = employee_by_role(s, name=to.name, role=to.role or "")
        _update_assignments_for_timeoff(
            s,
            employee=emp,
            start=to.from_date,
            end=to.to_date,
            approved=approved,
            timeoff=to,
            exclude_id=None if approved else to.id,
        )
        s.commit()
        week = s.scalar(select(Week).where(Week.start_date == date(2025, 9, 18)))
        if week:
            (
                counts,
                missing,
                required,
                variant_counts,
                shuttle_missing,
                shuttle_required,
                shuttle_counts,
                bb_missing,
                bb_required,
                bb_counts,
                fd_duplicates,
                maintenance_missing,
                maintenance_required,
                maintenance_counts,
            ) = coverage_snapshot_db(week.id)
            double_booked = double_booked_snapshot(week.id)
        else:
            counts = {}
            missing = {}
            required = 0
            variant_counts = {}
            shuttle_missing = {}
            shuttle_required = 0
            shuttle_counts = {}
            bb_missing = {}
            bb_required = 0
            bb_counts = {}
            fd_duplicates = {}
            maintenance_missing = {}
            maintenance_required = 0
            maintenance_counts = {}
            double_booked = {}
        item = {
            "id": to.id,
            "name": to.name,
            "role": to.role,
            "from": to.from_date.isoformat(),
            "to": to.to_date.isoformat(),
            "approved": to.approved,
            "vacation": bool(getattr(to, "vacation", False)),
        }
        # Broadcast to live listeners
        _broadcast({
            "type": "timeoff",
            "op": "toggle",
            "item": {
                **item,
                "vacation": bool(getattr(to, "vacation", False)),
            },
            "counts": counts,
            "missing": missing,
            "required": required,
            "variant_counts": variant_counts,
            "shuttle_missing": shuttle_missing,
            "shuttle_required": shuttle_required,
            "shuttle_counts": shuttle_counts,
            "bb_missing": bb_missing,
            "bb_required": bb_required,
            "bb_counts": bb_counts,
            "fd_duplicates": fd_duplicates,
            "maintenance_missing": maintenance_missing,
            "maintenance_required": maintenance_required,
            "maintenance_counts": maintenance_counts,
            "double_booked": double_booked,
        })
    return jsonify({
        "ok": True,
        "item": item,
        "counts": counts,
        "missing": missing,
        "required": required,
        "variant_counts": variant_counts,
        "shuttle_missing": shuttle_missing,
        "shuttle_required": shuttle_required,
        "shuttle_counts": shuttle_counts,
        "bb_missing": bb_missing,
        "bb_required": bb_required,
        "bb_counts": bb_counts,
        "fd_duplicates": fd_duplicates,
        "maintenance_missing": maintenance_missing,
        "maintenance_required": maintenance_required,
        "maintenance_counts": maintenance_counts,
        "double_booked": double_booked,
    })


@app.route("/timeoff/vacation", methods=["POST"])
def toggle_timeoff_vacation():
    data = request.get_json(force=True)
    tid = data.get("id")
    vacation = bool(data.get("vacation"))
    with SessionLocal() as s:
        to = s.get(TimeOff, tid)
        if not to:
            return jsonify({"ok": False, "error": "Not found"}), 404
        to.vacation = vacation
        emp = employee_by_role(s, name=to.name, role=to.role or "")
        if to.approved:
            _update_assignments_for_timeoff(
                s,
                employee=emp,
                start=to.from_date,
                end=to.to_date,
                approved=True,
                timeoff=to,
            )
        s.commit()

        response_item = {
            "id": to.id,
            "name": to.name,
            "role": to.role,
            "from": to.from_date.isoformat(),
            "to": to.to_date.isoformat(),
            "approved": bool(to.approved),
            "vacation": bool(to.vacation),
        }

        week = s.scalar(select(Week).where(Week.start_date == date(2025, 9, 18)))
        if week:
            (
                counts,
                missing,
                required,
                variant_counts,
                shuttle_missing,
                shuttle_required,
                shuttle_counts,
                bb_missing,
                bb_required,
                bb_counts,
                fd_duplicates,
                maintenance_missing,
                maintenance_required,
                maintenance_counts,
            ) = coverage_snapshot_db(week.id)
            payload = {
                "type": "timeoff",
                "op": "vacation",
                "item": response_item,
                "counts": counts,
                "missing": missing,
                "required": required,
                "variant_counts": variant_counts,
                "shuttle_missing": shuttle_missing,
                "shuttle_required": shuttle_required,
                "shuttle_counts": shuttle_counts,
                "bb_missing": bb_missing,
                "bb_required": bb_required,
                "bb_counts": bb_counts,
                "fd_duplicates": fd_duplicates,
                "maintenance_missing": maintenance_missing,
                "maintenance_required": maintenance_required,
                "maintenance_counts": maintenance_counts,
                "double_booked": double_booked_snapshot(week.id),
            }
        else:
            payload = {
                "type": "timeoff",
                "op": "vacation",
                "item": response_item,
            }
        _broadcast(payload)

        return jsonify({
            "ok": True,
            "item": response_item,
        })


def generate_new_schedule_db(week_id: int):
    """Legacy: generate a single week (kept for reference)."""
    with SessionLocal() as s:
        fd_emp_ids = list(s.scalars(select(Employee.id).join(Section).where(Section.name == "Front Desk")))
        bb_emp_ids = list(s.scalars(select(Employee.id).join(Section).where(Section.name == "Breakfast Bar")))
        sh_emp_ids = list(s.scalars(select(Employee.id).join(Section).where(Section.name == "Shuttle")))
        wk = s.get(Week, week_id)
        section_names = {sec.id: sec.name for sec in s.scalars(select(Section))}

        def _available_for_day(eid: int, dte: date) -> bool:
            emp = s.get(Employee, eid)
            if not emp:
                return False
            role_name = section_names.get(emp.section_id, "")
            return not has_any_timeoff(emp.name, role_name, dte, s)

        # Reset all
        s.execute(update(Assignment).where(Assignment.week_id == week_id).values(value="Set"))
        s.commit()

        # Breakfast Bar: ensure one person at 5am, 6am, and 7am each day
        for d in daterange(wk.start_date, 7):
            pool = bb_emp_ids[:]
            shifts = ["5AM–12PM", "6AM–12PM", "7AM–12PM"]
            
            # Assign one person to each of the three morning shifts
            for shift in shifts:
                if pool:  # Only assign if there are available employees
                    # pick someone without time off request
                    eligible = [eid for eid in pool if _available_for_day(eid, d)]
                    if not eligible:
                        continue
                    eid = sample(eligible, k=1)[0]
                    pool.remove(eid)
                    a = s.scalar(select(Assignment).where(Assignment.week_id == week_id, Assignment.employee_id == eid, Assignment.date == d))
                    if a:
                        a.value = shift
            
            # Remaining employees get "Set" (no shift)
            for eid in pool:
                a = s.scalar(select(Assignment).where(Assignment.week_id == week_id, Assignment.employee_id == eid, Assignment.date == d))
                if a:
                    a.value = "Set"

        # Front Desk: ensure 2 per variant (AM/PM/Audit) per day
        for d in daterange(wk.start_date, 7):
            pool = fd_emp_ids[:]
            # AM
            am_eligible = [eid for eid in pool if _available_for_day(eid, d)]
            am = sample(am_eligible, k=min(2, len(am_eligible))) if am_eligible else []
            # Senior earlier: sort by known seniority order
            def _sen_rank_fd(eid: int) -> int:
                name = s.get(Employee, eid).name if s.get(Employee, eid) else ''
                # Exception: Ryan should be treated as least-preferred for earlier start
                if name == 'Ryan':
                    return 10000
                try:
                    return SENIORITY_ORDER.index(name)
                except ValueError:
                    return 9999
            am.sort(key=_sen_rank_fd)
            for i, eid in enumerate(am):
                pool.remove(eid)
                a = s.scalar(select(Assignment).where(Assignment.week_id == week_id, Assignment.employee_id == eid, Assignment.date == d))
                if a:
                    # If only Ryan was picked, give him the later stagger
                    if len(am) == 1 and (s.get(Employee, eid).name if s.get(Employee, eid) else '') == 'Ryan':
                        a.value = "AM (6:15AM–2:15PM)"
                    else:
                        a.value = "AM (6:00AM–2:00PM)" if i == 0 else "AM (6:15AM–2:15PM)"
            # PM
            pm_candidates = [eid for eid in pool if _available_for_day(eid, d)]
            pm = sample(pm_candidates, k=min(2, len(pm_candidates))) if pm_candidates else []
            pm.sort(key=_sen_rank_fd)
            for i, eid in enumerate(pm):
                pool.remove(eid)
                a = s.scalar(select(Assignment).where(Assignment.week_id == week_id, Assignment.employee_id == eid, Assignment.date == d))
                if a:
                    if len(pm) == 1 and (s.get(Employee, eid).name if s.get(Employee, eid) else '') == 'Ryan':
                        a.value = "PM (2:15PM–10:15PM)"
                    else:
                        a.value = "PM (2:00PM–10:00PM)" if i == 0 else "PM (2:15PM–10:15PM)"
            # Audit
            au_candidates = [eid for eid in pool if _available_for_day(eid, d)]
            au = sample(au_candidates, k=min(2, len(au_candidates))) if au_candidates else []
            au.sort(key=_sen_rank_fd)
            for i, eid in enumerate(au):
                pool.remove(eid)
                a = s.scalar(select(Assignment).where(Assignment.week_id == week_id, Assignment.employee_id == eid, Assignment.date == d))
                if a:
                    if len(au) == 1 and (s.get(Employee, eid).name if s.get(Employee, eid) else '') == 'Ryan':
                        a.value = "Audit (10:15PM–6:15AM)"
                    else:
                        a.value = "Audit (10:00PM–6:00AM)" if i == 0 else "Audit (10:15PM–6:15AM)"

        # Shuttle: one agent per variant per day (as available)
        for d in daterange(wk.start_date, 7):
            pool = sh_emp_ids[:]
            variants = [
                "AM (3:30AM–11:30AM)",
                "Midday (10:30AM–6:30PM)",
                "PM (5:30PM–1:30AM)",
                DEFAULT_CREW_SHIFT,
            ]
            for v in variants:
                if not pool:
                    break
                eligible = [eid for eid in pool if _available_for_day(eid, d)]
                if not eligible:
                    continue
                eid = sample(eligible, k=1)[0]
                pool.remove(eid)
                a = s.scalar(select(Assignment).where(Assignment.week_id == week_id, Assignment.employee_id == eid, Assignment.date == d))
                if a:
                    a.value = v

        # Ensure approved time off overrides any generated shifts
        sync_timeoff_to_assignments(week_id, s)

        s.commit()


def _template_slot_label(slot: int) -> str:
    return f"Template {slot}"


def _serialize_template_slot(slot: int, row: Optional[ScheduleTemplate]) -> dict:
    if row:
        saved_range = format_week_label(row.saved_week_start) if row.saved_week_start else None
        saved_week_start = row.saved_week_start.isoformat() if row.saved_week_start else None
        updated_label = row.updated_at.strftime("%b %d, %Y %I:%M %p") if row.updated_at else None
        updated_iso = row.updated_at.isoformat() if row.updated_at else None
        return {
            "slot": slot,
            "name": row.name or _template_slot_label(slot),
            "has_data": bool(row.payload),
            "saved_week_start": saved_week_start,
            "saved_week_label": saved_range,
            "updated_label": updated_label,
            "updated_at": updated_iso,
        }
    return {
        "slot": slot,
        "name": _template_slot_label(slot),
        "has_data": False,
        "saved_week_start": None,
        "saved_week_label": None,
        "updated_label": None,
        "updated_at": None,
    }


def _capture_template_payload(session: Session, week: Week) -> dict:
    assignments = session.scalars(select(Assignment).where(Assignment.week_id == week.id)).all()
    items: list[dict[str, Any]] = []
    for assignment in assignments:
        if assignment.date is None:
            continue
        day_offset = (assignment.date - week.start_date).days
        if day_offset < 0 or day_offset > 6:
            continue
        items.append(
            {
                "employee_id": assignment.employee_id,
                "day": day_offset,
                "value": assignment.value or "Set",
                "dismissed": int(getattr(assignment, "dismissed_timeoff", 0) or 0),
            }
        )
    return {
        "version": 1,
        "week_start": week.start_date.isoformat(),
        "captured_at": datetime.utcnow().isoformat() + "Z",
        "assignments": items,
    }


def _template_slots_info(session: Session, total_slots: int = TEMPLATE_SLOT_COUNT) -> list[dict]:
    rows = {row.slot: row for row in session.scalars(select(ScheduleTemplate))}
    return [_serialize_template_slot(idx, rows.get(idx)) for idx in range(1, total_slots + 1)]


def _apply_template_payload_to_week(payload: dict, week: Week, session: Session) -> int:
    assignments = payload.get("assignments")
    if not isinstance(assignments, list):
        return 0
    updated = 0
    for entry in assignments:
        try:
            employee_id = int(entry.get("employee_id"))
            day_offset = int(entry.get("day"))
        except (TypeError, ValueError):
            continue
        if day_offset < 0 or day_offset > 6:
            continue
        target_date = week.start_date + timedelta(days=day_offset)
        employee = session.get(Employee, employee_id)
        if not employee:
            continue
        assignment = session.scalar(
            select(Assignment).where(
                Assignment.week_id == week.id,
                Assignment.employee_id == employee_id,
                Assignment.date == target_date,
            )
        )
        if not assignment:
            assignment = Assignment(
                week_id=week.id,
                employee_id=employee_id,
                date=target_date,
                value="Set",
            )
            session.add(assignment)
        value = entry.get("value") or "Set"
        assignment.value = value
        dismissed_flag = bool(entry.get("dismissed"))
        if hasattr(assignment, "dismissed_timeoff"):
            assignment.dismissed_timeoff = dismissed_flag  # type: ignore[attr-defined]
        updated += 1
    session.commit()
    sync_timeoff_to_assignments(week.id, session)
    return updated


def _ensure_week_and_assignments(s: Session, start_d: date) -> Week:
    wk = s.scalar(select(Week).where(Week.start_date == start_d))
    if not wk:
        wk = Week(start_date=start_d)
        s.add(wk)
        s.flush()
    # Ensure assignments exist for all employees/days
    for emp in s.scalars(select(Employee)):
        for d in daterange(start_d, 7):
            exists = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == emp.id, Assignment.date == d))
            if not exists:
                s.add(Assignment(week_id=wk.id, employee_id=emp.id, date=d, value="Set"))
    s.commit()
    return wk


# ---- Simple SSE broadcaster for live updates ----
from queue import Queue
from threading import Lock

_listeners: list[Queue] = []
_listeners_lock = Lock()


def _broadcast(event: dict):
    data = json.dumps(event)
    with _listeners_lock:
        for q in list(_listeners):
            try:
                q.put(data, block=False)
            except Exception:
                # Best-effort; drop if enqueue fails
                pass


@app.route("/events")
def sse_events():
    q: Queue = Queue()
    with _listeners_lock:
        _listeners.append(q)

    def gen():
        try:
            while True:
                msg = q.get()
                yield f"data: {msg}\n\n"
        finally:
            with _listeners_lock:
                if q in _listeners:
                    _listeners.remove(q)

    headers = {
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",  # for some proxies
        "Content-Type": "text/event-stream",
        "Connection": "keep-alive",
    }
    return Response(stream_with_context(gen()), headers=headers)


def _build_availability_index(s: Session) -> dict[int, set[tuple[int, str]]]:
    idx: dict[int, set[tuple[int, str]]] = {}
    for ea in s.scalars(select(EmployeeAvailability)):
        idx.setdefault(ea.employee_id, set()).add((ea.day_of_week, ea.shift_label))
    return idx


def _fd_variant(label: str) -> str:
    if label.startswith("AM"):
        return "AM"
    if label.startswith("PM"):
        return "PM"
    if label.startswith("Audit"):
        return "Audit"
    return label


def _is_available(emp_id: int, role: str, shift_label: str, dte: date, avail_idx: dict[int, set[tuple[int, str]]]) -> bool:
    allowed = avail_idx.get(emp_id)
    # If no availability configured, treat as unavailable
    # This prevents scheduling employees who have not set any availability.
    if not allowed:
        return False
    dow = dte.weekday()  # 0=Mon
    if role == "Front Desk":
        key = (dow, _fd_variant(shift_label))
    else:
        key = (dow, shift_label)
    return key in allowed


def _pref_token(role: str, label: str) -> str:
    if role == "Front Desk":
        return _fd_variant(label)
    return label


def _effective_max_for_role(emp_info: dict[int, dict], eid: int, role: str) -> Optional[int]:
    """Return the max shifts/week for this employee in the context of a role.
    If employee-specific max is set, use it; otherwise default per role:
    - Front Desk: 5
    - Shuttle: 4
    - Breakfast Bar: no default cap (None)
    """
    maxw = emp_info[eid].get("max_per_week")
    if maxw is not None:
        return maxw
    if role == "Front Desk":
        return 5
    if role == "Shuttle":
        return 4
    return None


def generate_4_week_schedule(start_week_id: int):
    """Generate a 4-week schedule starting at the given week.
    Respects EmployeeAvailability; leaves unfilled if no available employees.
    Also reapplies approved time off per week.
    """
    with SessionLocal() as s:
        base_week = s.get(Week, start_week_id)
        if not base_week:
            return

        # Simple file logger to capture scheduling decisions
        def _log(msg: str):
            try:
                with open('log.txt', 'a', encoding='utf-8') as _f:
                    _f.write(msg.rstrip() + "\n")
            except Exception:
                pass

        # Truncate previous log
        try:
            open('log.txt', 'w').close()
        except Exception:
            pass
        _log(f"Generate 4-week schedule starting week_id={start_week_id} ({base_week.start_date.isoformat()})")

        # Precompute availability and role mappings
        avail_idx = _build_availability_index(s)
        emp_info: dict[int, dict] = {}
        emp_name: dict[int, str] = {}
        for e in s.scalars(select(Employee)):
            sec = s.get(Section, e.section_id)
            emp_info[e.id] = {
                "role": (sec.name if sec else ""),
                "seniority": e.seniority or 0,
                "preferred_shift": e.preferred_shift or None,
                "pref_per_week": e.preferred_shifts_per_week if (e.preferred_shifts_per_week is not None) else None,
                "max_per_week": e.max_shifts_per_week if (e.max_shifts_per_week is not None) else None,
            }
            emp_name[e.id] = e.name

        # Gather employees by section including secondary roles
        def employees_for_section(name: str) -> list[int]:
            sec = s.scalar(select(Section).where(Section.name == name))
            if not sec:
                return []
            primary_emps = list(
                s.scalars(
                    select(Employee)
                    .where(Employee.section_id == sec.id)
                    .order_by(Employee.sort_order.is_(None), Employee.sort_order, Employee.name)
                )
            )
            secondary_ids = [row[0] for row in s.execute(select(EmployeeRole.employee_id).where(EmployeeRole.section_id == sec.id)).all()]
            secondary_emps = (
                list(
                    s.scalars(
                        select(Employee)
                        .where(Employee.id.in_(secondary_ids))
                        .order_by(Employee.sort_order.is_(None), Employee.sort_order, Employee.name)
                    )
                )
                if secondary_ids
                else []
            )
            seen = {e.id for e in primary_emps}
            ordered = primary_emps + [e for e in secondary_emps if e.id not in seen]
            return [e.id for e in ordered]

        fd_emp_ids = employees_for_section("Front Desk")
        bb_emp_ids = employees_for_section("Breakfast Bar")
        sh_emp_ids = employees_for_section("Shuttle")
        maint_emp_ids = employees_for_section("Maintenance")
        _log(f"FD-capable: {[emp_name.get(i,'?') for i in fd_emp_ids]}")
        _log(f"BB-capable: {[emp_name.get(i,'?') for i in bb_emp_ids]}")
        _log(f"SH-capable: {[emp_name.get(i,'?') for i in sh_emp_ids]}")
        _log(f"MA-capable: {[emp_name.get(i,'?') for i in maint_emp_ids]}")

        for week_offset in range(4):
            start_d = base_week.start_date + timedelta(days=7 * week_offset)
            wk = _ensure_week_and_assignments(s, start_d)

            # Reset all to Set and clear dismissed flags (if column exists)
            if hasattr(Assignment, 'dismissed_timeoff'):
                s.execute(update(Assignment).where(Assignment.week_id == wk.id).values(value="Set", dismissed_timeoff=0))
            else:
                s.execute(update(Assignment).where(Assignment.week_id == wk.id).values(value="Set"))
            s.commit()

            # Track assigned count per employee within this week (excludes TIME OFF)
            assigned_counts: dict[int, int] = {eid: 0 for eid in emp_info.keys()}

            # Track last assigned token per employee per role to prefer consistency within the week
            last_token_per_role: dict[tuple[int, str], Optional[str]] = {}

            def rest_ok(eid: int, dte: date, role: str, label: str) -> bool:
                # Apply simple rest constraints across consecutive days, even over week boundaries
                prev_date = dte - timedelta(days=1)
                a_prev = s.scalar(select(Assignment).where(Assignment.employee_id == eid, Assignment.date == prev_date))
                if not a_prev or not a_prev.value or a_prev.value in NEUTRAL_ASSIGNMENT_VALUES:
                    return True
                # No AM after a previous-day PM (short 8h rest)
                if role == "Front Desk" and label.startswith("AM") and a_prev.value.startswith("PM"):
                    return False
                # No PM after a previous-night Audit (8h rest window)
                if role == "Front Desk" and label.startswith("PM") and a_prev.value.startswith("Audit"):
                    return False
                # Shuttle rest rules
                if role == "Shuttle":
                    # No AM after a previous-day late shift (PM or Crew ends after midnight)
                    prev_val = a_prev.value or ""
                    prev_was_crew = prev_val.startswith("Crew") or is_suggested_crew_label(prev_val)
                    if label.startswith("AM") and (prev_val.startswith("PM") or prev_was_crew):
                        return False
                    # Be cautious with Midday after Crew (only ~8h45m rest)
                    if label.startswith("Midday") and prev_was_crew:
                        return False
                return True

            def pick_best(candidates: list[int], role: str, label: str) -> Optional[int]:
                # Filter out those exceeding max per week
                filtered = []
                for eid in candidates:
                    maxw = _effective_max_for_role(emp_info, eid, role)
                    if maxw is not None and assigned_counts[eid] >= maxw:
                        _log(f"{current_loop_date} {role} {label}: drop {emp_name.get(eid,'?')} (max/week reached {maxw})")
                        continue
                    # Rest constraints
                    # Determine actual date in outer calling context by closure hack: rely on current_loop_date variable
                    if not rest_ok(eid, current_loop_date, role, label):
                        _log(f"{current_loop_date} {role} {label}: drop {emp_name.get(eid,'?')} (rest rule)")
                        continue
                    filtered.append(eid)
                if not filtered:
                    _log(f"{current_loop_date} {role} {label}: no candidates after filters")
                    return None
                def key_fn(eid: int):
                    under_pref = False
                    prefw = emp_info[eid]["pref_per_week"]
                    if prefw is not None:
                        under_pref = assigned_counts[eid] < prefw
                    # Compare this employee's own preferred shift against the current token
                    want = emp_info[eid]["preferred_shift"]
                    preferred_match = (want is not None and _pref_token(role, label) == want)
                    # Prefer consistency with last assigned token within the week for this role
                    last_tok = last_token_per_role.get((eid, role))
                    consistent = (last_tok is not None and last_tok == _pref_token(role, label))
                    return (
                        0 if under_pref else 1,
                        0 if preferred_match else 1,
                        0 if consistent else 1,
                        -(emp_info[eid]["seniority"] or 0),
                        assigned_counts[eid],
                    )
                ranked = sorted(filtered, key=key_fn)
                _log(f"{current_loop_date} {role} {label}: ranked {[emp_name.get(e,'?') for e in ranked]}")
                best = ranked[0]
                return best

            # Helper to mark if candidate dropped solely due to time off
            def _mark_dismissed(eid: int, dte: date):
                if not hasattr(Assignment, 'dismissed_timeoff'):
                    return
                a = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == eid, Assignment.date == dte))
                if a and hasattr(a, 'dismissed_timeoff'):
                    a.dismissed_timeoff = 1

            # Front Desk: 2 per AM/PM/Audit per day if available (prioritize FD first)
            for d in daterange(wk.start_date, 7):
                current_loop_date = d
                pool = fd_emp_ids[:]
                for variant, opts in (
                    ("AM", ["AM (6:00AM–2:00PM)", "AM (6:15AM–2:15PM)"]),
                    ("PM", ["PM (2:00PM–10:00PM)", "PM (2:15PM–10:15PM)"]),
                    ("Audit", ["Audit (10:00PM–6:00AM)", "Audit (10:15PM–6:15AM)"]),
                ):
                    # Choose up to 2 employees available and not already assigned that day
                    cand = []
                    _log(f"{d} FD {variant}: start pool {[emp_name.get(i,'?') for i in pool]}")
                    for eid in pool:
                        a = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == eid, Assignment.date == d))
                        if not a or a.value != "Set":
                            _log(f"{d} FD {variant}: drop {emp_name.get(eid,'?')} (already assigned {a.value if a else 'None'})")
                            continue
                        if not _is_available(eid, "Front Desk", variant, d, avail_idx):
                            _log(f"{d} FD {variant}: drop {emp_name.get(eid,'?')} (not available)")
                            continue
                        if not rest_ok(eid, d, "Front Desk", variant):
                            _log(f"{d} FD {variant}: drop {emp_name.get(eid,'?')} (rest rule)")
                            continue
                        # Skip anyone with any time off request on this date
                        if has_any_timeoff(emp_name.get(eid, ''), emp_info.get(eid, {}).get("role", ""), d, s):
                            _log(f"{d} FD {variant}: drop {emp_name.get(eid,'?')} (time off request)")
                            _mark_dismissed(eid, d)
                            continue
                        if _effective_max_for_role(emp_info, eid, "Front Desk") is not None and assigned_counts[eid] >= _effective_max_for_role(emp_info, eid, "Front Desk"):
                            _log(f"{d} FD {variant}: drop {emp_name.get(eid,'?')} (max/week reached)")
                            continue
                        # candidate accepted
                        cand.append(eid)
                    _log(f"{d} FD {variant}: candidates {[emp_name.get(i,'?') for i in cand]}")
                    if not cand:
                        continue
                    picks: list[int] = []
                    for slot_i in range(2):
                        pick = pick_best([e for e in cand if e not in picks], "Front Desk", variant)
                        if pick is None:
                            break
                        picks.append(pick)
                    # Senior earlier: ensure earlier stagger goes to more senior
                    def _seniority_rank(eid: int) -> int:
                        name = emp_name.get(eid, '')
                        # Exception: Ryan should get the later stagger, so sort him last
                        if name == 'Ryan':
                            return 10000
                        try:
                            return SENIORITY_ORDER.index(name)
                        except ValueError:
                            return 9999
                    picks.sort(key=_seniority_rank)
                    for i, eid in enumerate(picks):
                        pool.remove(eid)
                        # If only Ryan was picked for this variant, give him the later stagger
                        if len(picks) == 1 and emp_name.get(eid, '') == 'Ryan' and len(opts) > 1:
                            label = opts[1]
                        else:
                            label = opts[i] if i < len(opts) else opts[-1]
                        a = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == eid, Assignment.date == d))
                        if a:
                            a.value = label
                            assigned_counts[eid] += 1
                            last_token_per_role[(eid, "Front Desk")] = _pref_token("Front Desk", variant)
                            _log(f"{d} FD assign {label} -> {emp_name.get(eid,'?')}")

            # After FD assignments, strictly reserve FD-capable employees when FD coverage is still missing
            reserved_fd_by_date: dict[date, set[int]] = {}
            def fd_missing_on(day: date) -> bool:
                fd_counts = {"AM": 0, "PM": 0, "Audit": 0}
                for eid in fd_emp_ids:
                    a = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == eid, Assignment.date == day))
                    if not a or not a.value or a.value in NEUTRAL_ASSIGNMENT_VALUES:
                        continue
                    if a.value.startswith("AM"):
                        fd_counts["AM"] += 1
                    elif a.value.startswith("PM"):
                        fd_counts["PM"] += 1
                    elif a.value.startswith("Audit"):
                        fd_counts["Audit"] += 1
                missing = any(fd_counts[v] < 2 for v in ("AM", "PM", "Audit"))
                if missing:
                    _log(f"{day} FD missing counts={fd_counts}")
                return missing
            for d in daterange(wk.start_date, 7):
                # Count current FD coverage among FD-capable employees
                fd_counts = {"AM": 0, "PM": 0, "Audit": 0}
                for eid in fd_emp_ids:
                    a = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == eid, Assignment.date == d))
                    if not a or not a.value or a.value in NEUTRAL_ASSIGNMENT_VALUES:
                        continue
                    if a.value.startswith("AM"):
                        fd_counts["AM"] += 1
                    elif a.value.startswith("PM"):
                        fd_counts["PM"] += 1
                    elif a.value.startswith("Audit"):
                        fd_counts["Audit"] += 1
                if any(fd_counts[v] < 2 for v in ("AM", "PM", "Audit")):
                    # Reserve all FD-capable employees who are still unassigned (Set) on this day
                    reserve: set[int] = set()
                    for eid in fd_emp_ids:
                        a = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == eid, Assignment.date == d))
                        if a and a.value == "Set":
                            reserve.add(eid)
                    if reserve:
                        reserved_fd_by_date[d] = reserve
                        _log(f"{d} FD reserve {[emp_name.get(i,'?') for i in sorted(reserve)]}")

            # Maintenance: one agent per day if available
            for d in daterange(wk.start_date, 7):
                current_loop_date = d
                candidates: list[int] = []
                for eid in maint_emp_ids:
                    a = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == eid, Assignment.date == d))
                    if not a or a.value != "Set":
                        _log(f"{d} MA: drop {emp_name.get(eid,'?')} (assigned {a.value if a else 'None'})")
                        continue
                    if has_any_timeoff(emp_name.get(eid, ''), emp_info.get(eid, {}).get("role", ""), d, s):
                        _log(f"{d} MA: drop {emp_name.get(eid,'?')} (time off request)")
                        _mark_dismissed(eid, d)
                        continue
                    if not _is_available(eid, "Maintenance", "8AM–4:30PM", d, avail_idx):
                        _log(f"{d} MA: drop {emp_name.get(eid,'?')} (not available)")
                        continue
                    if not rest_ok(eid, d, "Maintenance", "8AM–4:30PM"):
                        _log(f"{d} MA: drop {emp_name.get(eid,'?')} (rest rule)")
                        continue
                    candidates.append(eid)
                pick = pick_best(candidates, "Maintenance", "8AM–4:30PM")
                if pick is None:
                    continue
                a = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == pick, Assignment.date == d))
                if a:
                    a.value = "8AM–4:30PM"
                    assigned_counts[pick] += 1
                    last_token_per_role[(pick, "Maintenance")] = _pref_token("Maintenance", "8AM–4:30PM")
                    _log(f"{d} MA assign 8AM–4:30PM -> {emp_name.get(pick,'?')}")

            # Breakfast Bar: one per 5/6/7am per day if available (after FD; skip reserved FD candidates and prefer non-FD when FD is missing)
            for d in daterange(wk.start_date, 7):
                current_loop_date = d
                pool = bb_emp_ids[:]
                for shift in ["5AM–12PM", "6AM–12PM", "7AM–12PM"]:
                    candidates = []
                    for eid in pool:
                        # Respect FD reservation for this day
                        if d in reserved_fd_by_date and eid in reserved_fd_by_date[d]:
                            _log(f"{d} BB {shift}: skip {emp_name.get(eid,'?')} (reserved for FD)")
                            continue
                        # If FD is missing today, avoid using FD-capable people in BB
                        if fd_missing_on(d) and eid in fd_emp_ids:
                            _log(f"{d} BB {shift}: skip {emp_name.get(eid,'?')} (FD missing; FD-capable)")
                            continue
                        a = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == eid, Assignment.date == d))
                        if not a or a.value != "Set":
                            _log(f"{d} BB {shift}: drop {emp_name.get(eid,'?')} (already assigned {a.value if a else 'None'})")
                            continue
                        if has_any_timeoff(emp_name.get(eid, ''), emp_info.get(eid, {}).get("role", ""), d, s):
                            _log(f"{d} BB {shift}: drop {emp_name.get(eid,'?')} (time off request)")
                            # Mark dismissed only if they'd otherwise pass availability/rest for this shift
                            if _is_available(eid, "Breakfast Bar", shift, d, avail_idx) and rest_ok(eid, d, "Breakfast Bar", shift):
                                _mark_dismissed(eid, d)
                            continue
                        if _is_available(eid, "Breakfast Bar", shift, d, avail_idx) and rest_ok(eid, d, "Breakfast Bar", shift):
                            candidates.append(eid)
                        else:
                            _log(f"{d} BB {shift}: drop {emp_name.get(eid,'?')} (not available/rest)")
                    pick = pick_best(candidates, "Breakfast Bar", shift)
                    if pick is None:
                        continue
                    pool.remove(pick)
                    a = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == pick, Assignment.date == d))
                    if a:
                        a.value = shift
                        assigned_counts[pick] += 1
                        last_token_per_role[(pick, "Breakfast Bar")] = _pref_token("Breakfast Bar", shift)
                        _log(f"{d} BB assign {shift} -> {emp_name.get(pick,'?')}")

            # Shuttle: one per variant per day if available (skip reserved FD candidates and prefer non-FD when FD is missing)
            for d in daterange(wk.start_date, 7):
                current_loop_date = d
                pool = sh_emp_ids[:]
                variants = [
                    "AM (3:30AM–11:30AM)",
                    "Midday (10:30AM–6:30PM)",
                    "PM (5:30PM–1:30AM)",
                    DEFAULT_CREW_SHIFT,
                ]
                for v in variants:
                    cand = []
                    for eid in pool:
                        # Respect FD reservation for this day
                        if d in reserved_fd_by_date and eid in reserved_fd_by_date[d]:
                            _log(f"{d} SH {v}: skip {emp_name.get(eid,'?')} (reserved for FD)")
                            continue
                        # If FD is missing today, avoid using FD-capable people in Shuttle
                        if fd_missing_on(d) and eid in fd_emp_ids:
                            _log(f"{d} SH {v}: skip {emp_name.get(eid,'?')} (FD missing; FD-capable)")
                            continue
                        a = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == eid, Assignment.date == d))
                        if not a or a.value != "Set":
                            _log(f"{d} SH {v}: drop {emp_name.get(eid,'?')} (already assigned {a.value if a else 'None'})")
                            continue
                        if has_any_timeoff(emp_name.get(eid, ''), emp_info.get(eid, {}).get("role", ""), d, s):
                            _log(f"{d} SH {v}: drop {emp_name.get(eid,'?')} (time off request)")
                            if rest_ok(eid, d, "Shuttle", v):
                                _mark_dismissed(eid, d)
                            continue
                        if _is_available(eid, "Shuttle", v, d, avail_idx) and rest_ok(eid, d, "Shuttle", v):
                            cand.append(eid)
                        else:
                            _log(f"{d} SH {v}: drop {emp_name.get(eid,'?')} (not available/rest)")
                    pick = pick_best(cand, "Shuttle", v)
                    if pick is None:
                        continue
                    pool.remove(pick)
                    a = s.scalar(select(Assignment).where(Assignment.week_id == wk.id, Assignment.employee_id == pick, Assignment.date == d))
                    if a:
                        a.value = v
                        assigned_counts[pick] += 1
                        last_token_per_role[(pick, "Shuttle")] = _pref_token("Shuttle", v)
                        _log(f"{d} SH assign {v} -> {emp_name.get(pick,'?')}")

            # Ensure approved time off overrides any generated shifts for this week
            sync_timeoff_to_assignments(wk.id, s)

        s.commit()


@app.route("/generate", methods=["POST", "GET"]) 
def generate():
    with SessionLocal() as s:
        week = s.scalar(select(Week).where(Week.start_date == date(2025, 9, 18)))
        week_id = week.id
    generate_4_week_schedule(week_id)
    return redirect(url_for("index"))


@app.route("/week/<int:week_id>/generate", methods=["POST"]) 
def generate_week(week_id: int):
    generate_4_week_schedule(week_id)
    return redirect(url_for("view_week", week_id=week_id))


def _delete_period_with_undo(s: Session, week: Week) -> Optional[str]:
    if not week:
        return None

    period_start, period_end = four_week_period_bounds(week.start_date)
    period_weeks = list(
        s.scalars(
            select(Week).where(Week.start_date >= period_start, Week.start_date <= period_end).order_by(Week.start_date)
        )
    )
    if not period_weeks:
        return None

    weeks_payload: List[dict] = []
    for pw in period_weeks:
        assignments_payload: List[dict] = []
        for a in s.scalars(select(Assignment).where(Assignment.week_id == pw.id)):
            assignment_info = {
                "employee_id": a.employee_id,
                "date": a.date.isoformat(),
                "value": a.value,
            }
            if hasattr(Assignment, "dismissed_timeoff"):
                assignment_info["dismissed"] = int(getattr(a, "dismissed_timeoff", 0) or 0)
            assignments_payload.append(assignment_info)
        weeks_payload.append({
            "start_date": pw.start_date.isoformat(),
            "assignments": assignments_payload,
        })

    _prune_expired_period_undos()
    expires_at = time.time() + UNDO_DELETE_SECONDS
    token = uuid.uuid4().hex
    label = format_four_week_label(period_start, period_end)
    _pending_period_undos[token] = {
        "expires": expires_at,
        "label": label,
        "weeks": weeks_payload,
    }

    try:
        for pw in period_weeks:
            if pw.start_date == FOUR_WEEK_BASELINE:
                values = {"value": "Set"}
                if hasattr(Assignment, "dismissed_timeoff"):
                    values["dismissed_timeoff"] = 0
                s.execute(update(Assignment).where(Assignment.week_id == pw.id).values(**values))
                continue
            s.execute(delete(Assignment).where(Assignment.week_id == pw.id))
            s.delete(pw)
        s.commit()
    except Exception:
        _pending_period_undos.pop(token, None)
        s.rollback()
        raise

    return token


@app.route("/delete", methods=["POST"]) 
def delete_schedule():
    with SessionLocal() as s:
        week = s.scalar(select(Week).where(Week.start_date == FOUR_WEEK_BASELINE))
        token = _delete_period_with_undo(s, week) if week else None
    redirect_args = {}
    if token:
        redirect_args["undo"] = token
    return redirect(url_for("schedules_page", **redirect_args))


@app.route("/week/<int:week_id>/delete", methods=["POST"]) 
def delete_week_schedule(week_id: int):
    with SessionLocal() as s:
        wk = s.get(Week, week_id)
        if not wk:
            return redirect(url_for("schedules_page"))

        token = _delete_period_with_undo(s, wk)
    redirect_args = {}
    if token:
        redirect_args["undo"] = token
    return redirect(url_for("schedules_page", **redirect_args))


@app.route("/undo-delete/<token>", methods=["POST"])
def undo_delete_period(token: str):
    _prune_expired_period_undos()
    payload = _pending_period_undos.get(token)
    if not payload or payload.get("expires", 0) <= time.time():
        _pending_period_undos.pop(token, None)
        return redirect(url_for("schedules_page", undo_status="expired"))

    weeks_payload = payload.get("weeks", [])
    with SessionLocal() as s:
        for week_info in weeks_payload:
            start = date.fromisoformat(week_info["start_date"])
            week = s.scalar(select(Week).where(Week.start_date == start))
            if not week:
                week = Week(start_date=start)
                s.add(week)
                s.flush()
            s.execute(delete(Assignment).where(Assignment.week_id == week.id))
            for assignment_info in week_info.get("assignments", []):
                assign_date = date.fromisoformat(assignment_info["date"])
                assignment = Assignment(
                    week_id=week.id,
                    employee_id=assignment_info["employee_id"],
                    date=assign_date,
                    value=assignment_info["value"],
                )
                if hasattr(Assignment, "dismissed_timeoff"):
                    assignment.dismissed_timeoff = assignment_info.get("dismissed", 0) or 0
                s.add(assignment)
        s.commit()

    label = payload.get("label")
    _pending_period_undos.pop(token, None)
    redirect_args = {"undo_status": "restored"}
    if label:
        redirect_args["label"] = label
    return redirect(url_for("schedules_page", **redirect_args))


@app.route("/week/<int:week_id>/prev")
def prev_week(week_id: int):
    """Navigate to the previous week"""
    with SessionLocal() as s:
        current_week = s.get(Week, week_id)
        if not current_week:
            return redirect(url_for("index"))
        
        # Calculate previous week start date (7 days earlier)
        prev_start_date = current_week.start_date - timedelta(days=7)
        
        # Check if previous week exists, create if not
        prev_week = s.scalar(select(Week).where(Week.start_date == prev_start_date))
        if not prev_week:
            prev_week = Week(start_date=prev_start_date)
            s.add(prev_week)
            s.flush()
            
            # Create assignments for all employees for the new week
            for emp in s.scalars(select(Employee)):
                for d in daterange(prev_start_date, 7):
                    s.add(Assignment(week_id=prev_week.id, employee_id=emp.id, date=d, value="Set"))
            s.commit()
            
            # Sync any approved time off for this week
            sync_timeoff_to_assignments(prev_week.id, s)
        
        return redirect(url_for("view_week", week_id=prev_week.id))


@app.route("/week/<int:week_id>/next")
def next_week(week_id: int):
    """Navigate to the next week"""
    with SessionLocal() as s:
        current_week = s.get(Week, week_id)
        if not current_week:
            return redirect(url_for("index"))
        
        # Calculate next week start date (7 days later)
        next_start_date = current_week.start_date + timedelta(days=7)
        
        # Check if next week exists, create if not
        next_week = s.scalar(select(Week).where(Week.start_date == next_start_date))
        if not next_week:
            next_week = Week(start_date=next_start_date)
            s.add(next_week)
            s.flush()
            
            # Create assignments for all employees for the new week
            for emp in s.scalars(select(Employee)):
                for d in daterange(next_start_date, 7):
                    s.add(Assignment(week_id=next_week.id, employee_id=emp.id, date=d, value="Set"))
            s.commit()
            
            # Sync any approved time off for this week
            sync_timeoff_to_assignments(next_week.id, s)
        
        return redirect(url_for("view_week", week_id=next_week.id))


@app.route("/schedule-template/upload", methods=["POST"])
def upload_schedule_template():
    upload = request.files.get("template")
    week_id = request.form.get("week_id", type=int)
    redirect_path = _sanitize_redirect_target(request.form.get("redirect_path"))
    target_url = (
        redirect_path
        or (url_for("view_week", week_id=week_id) if week_id else url_for("index"))
    )

    def fail(message: str):
        return _redirect_with_template_status(target_url, "error", message)

    if upload is None or not upload.filename:
        return fail("Select a .xlsx file to upload.")

    suffix = Path(upload.filename).suffix.lower()
    if suffix not in SCHEDULE_TEMPLATE_ALLOWED_SUFFIXES:
        return fail("Only .xlsx files are supported.")

    payload = upload.read()
    if not payload:
        return fail("The uploaded file is empty.")

    try:
        wb = load_workbook(io.BytesIO(payload))
        wb.close()
    except Exception:
        return fail("Unable to read that Excel file. Please upload a valid .xlsx template.")

    archive_path: Optional[Path] = None
    if SCHEDULE_TEMPLATE_FILENAME.exists():
        try:
            SCHEDULE_TEMPLATE_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
            archive_path = SCHEDULE_TEMPLATE_ARCHIVE_DIR / f"ScheduleTemplate-{timestamp}.xlsx"
            shutil.move(str(SCHEDULE_TEMPLATE_FILENAME), archive_path)
        except Exception as exc:
            app.logger.error("Unable to archive schedule template: %s", exc)
            return fail("Unable to archive the existing template. Please try again.")

    try:
        with open(SCHEDULE_TEMPLATE_FILENAME, "wb") as f:
            f.write(payload)
    except Exception as exc:
        app.logger.error("Unable to save schedule template: %s", exc)
        if archive_path and archive_path.exists():
            try:
                shutil.move(str(archive_path), SCHEDULE_TEMPLATE_FILENAME)
            except Exception:
                app.logger.error("Unable to restore previous schedule template after failure.")
        return fail("Unable to save the new template. Please try again.")

    success_message = "Schedule template updated successfully."
    if archive_path:
        success_message = f"Template replaced. Previous version saved to {archive_path.name}."
    return _redirect_with_template_status(target_url, "success", success_message)


@app.route("/schedule-template/download")
def download_schedule_template():
    if not SCHEDULE_TEMPLATE_FILENAME.exists():
        return Response("Schedule template not found.", status=404, mimetype="text/plain")
    safe_name = f"ScheduleTemplate-{date.today().isoformat()}.xlsx"
    return send_file(
        SCHEDULE_TEMPLATE_FILENAME,
        as_attachment=True,
        download_name=safe_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/excel/<int:week_id>")
def export_schedule_excel(week_id: int):
    """Export schedule to Excel file using the existing template"""
    with SessionLocal() as s:
        week = s.get(Week, week_id)
        if not week:
            return jsonify({"error": "Week not found"}), 404
        
        # Check if template exists
        template_path = SCHEDULE_TEMPLATE_FILENAME
        if not template_path.exists():
            return jsonify({"error": "ScheduleTemplate.xlsx not found"}), 404
        
        # Get schedule data
        ctx = build_week_context(week_id)
        dates = ctx["week"]["dates"]
        sections = ctx["week"]["sections"]
        occupancy_values = (ctx.get("occupancy") or {}).get("values") or {}
        aircrew_ctx = ctx.get("aircrew") or {}
        carriers = aircrew_ctx.get("carriers") or []
        carrier_arrivals = aircrew_ctx.get("arrivals") or {}
        
        # Load the template
        wb = load_workbook(template_path)
        ws = wb.active

        # Update the sheet title with a safe, short title (Excel max 31 chars)
        def safe_sheet_title(title: str) -> str:
            invalid = set('[]:*?/\\')
            cleaned = ''.join(ch for ch in title if ch not in invalid)
            return cleaned[:31]

        ws.title = safe_sheet_title(format_week_label(week.start_date))
        
        # Update the dates in row 3 (columns E-K)
        for i, date_info in enumerate(dates):
            col_letter = get_column_letter(5 + i)  # Start from column E (5)
            cell_ref = f'{col_letter}3'
            ws[cell_ref] = week.start_date + timedelta(days=i)
            try:
                # Display as e.g., 21-Aug
                ws[cell_ref].number_format = 'dd-mmm'
            except Exception:
                pass
        
        # Update day headers in row 4 (columns E-K)
        day_headers = ["THU", "FRI", "SAT", "SUN", "MON", "TUE", "WED"]
        for i, day in enumerate(day_headers):
            col_letter = get_column_letter(5 + i)  # Start from column E (5)
            ws[f'{col_letter}4'] = day
        
        # Define employee row mappings based on template structure
        # Front Desk: rows 5-22 (names in column D)
        # Shuttle: rows 24-35 (names in column D)
        # Breakfast Bar: rows 37-42 (names in column D)
        # Maintenance: starts at the row whose column A label is "Maintenance"
        
        # Create mapping of employee (name, section) pairs to their row numbers
        employee_rows: dict[tuple[str, str], dict[str, Any]] = {}
        section_blocks: dict[str, dict[str, Any]] = {}
        import re as _re_names

        def _primary_name_from_cell(raw: str) -> Optional[str]:
            """Extract the primary employee name from a template name cell.
            Handles cases like "Sara and TBD", "Sara/TBD", "Sara & TBD" by
            returning just "SARA" and ignoring placeholders like "TBD" or "-".
            Returns uppercase name or None if it's a placeholder or helper row (e.g., **OCC).
            """
            if not raw or not isinstance(raw, str):
                return None
            raw_stripped = raw.strip()
            # Template helper rows (occupancy, carrier labels) start with '**'
            if raw_stripped.startswith("**"):
                return None
            s = raw_stripped.replace("*", "").upper()
            if not s or s in {"-", "TBD"}:
                return None
            # Split on common connectors and pick the first non-placeholder token
            parts = _re_names.split(r"\s*(?:AND|/|&|\+)\s*", s)
            for p in parts:
                token = p.strip()
                if token and token not in {"-", "TBD"}:
                    return token
            return None

        def _is_blank_slot(value: Any) -> bool:
            if value is None:
                return True
            if not isinstance(value, str):
                return False
            stripped = value.strip()
            return stripped == "" or stripped in {"-", "—"}

        section_aliases = {
            "front desk": "Front Desk",
            "frontdesk": "Front Desk",
            "shuttle": "Shuttle",
            "shuttle drivers": "Shuttle",
            "breakfast bar": "Breakfast Bar",
            "maintenance": "Maintenance",
        }
        tracked_sections = set(section_aliases.values())
        current_section: Optional[str] = None

        def _section_block(name: str) -> dict[str, Any]:
            return section_blocks.setdefault(
                name,
                {
                    "blank_rows": [],
                    "end_row": None,
                    "template_row": None,
                },
            )

        for row in range(1, ws.max_row + 1):
            label_val = ws[f'A{row}'].value
            if isinstance(label_val, str):
                normalized_label = label_val.strip().lower()
                if normalized_label in section_aliases:
                    # close out previous section boundaries if necessary
                    if current_section and current_section != section_aliases[normalized_label]:
                        block = _section_block(current_section)
                        if block.get("end_row") is None:
                            block["end_row"] = row
                    current_section = section_aliases[normalized_label]
                elif normalized_label:
                    if current_section:
                        block = _section_block(current_section)
                        if block.get("end_row") is None:
                            block["end_row"] = row
                    current_section = None
            if current_section not in tracked_sections:
                continue
            block = _section_block(current_section)
            name_cell = ws[f'D{row}']
            primary = _primary_name_from_cell(name_cell.value)
            if primary:
                employee_rows[(primary, current_section)] = {"row": row, "section": current_section}
                if block.get("template_row") is None:
                    block["template_row"] = row
            elif _is_blank_slot(name_cell.value):
                block["blank_rows"].append(row)
        # Ensure end boundaries for any trailing sections
        for block in section_blocks.values():
            if block.get("end_row") is None:
                block["end_row"] = ws.max_row + 1

        def _clone_row_style(src_row: int, dest_row: int) -> None:
            if src_row < 1 or dest_row < 1 or src_row > ws.max_row:
                return
            for col_idx in range(4, min(ws.max_column, 11) + 1):
                src_cell = ws.cell(row=src_row, column=col_idx)
                dest_cell = ws.cell(row=dest_row, column=col_idx)
                dest_cell.value = None
                if src_cell.has_style:
                    dest_cell.font = copy_style(src_cell.font)
                    dest_cell.border = copy_style(src_cell.border)
                    dest_cell.fill = copy_style(src_cell.fill)
                    dest_cell.number_format = src_cell.number_format
                    dest_cell.alignment = copy_style(src_cell.alignment)
            src_dim = ws.row_dimensions.get(src_row)
            if src_dim and src_dim.height:
                ws.row_dimensions[dest_row].height = src_dim.height

        def _ensure_employee_row(section_name: str, employee_key: str, display_name: str) -> Optional[int]:
            existing = employee_rows.get((employee_key, section_name))
            if existing and existing["section"] == section_name:
                return existing["row"]

            block = section_blocks.get(section_name)
            if not block:
                return None

            if block["blank_rows"]:
                row_num = block["blank_rows"].pop(0)
            else:
                insert_at = block.get("end_row") or (ws.max_row + 1)
                ws.insert_rows(insert_at)
                template_row = block.get("template_row")
                reference_row = template_row or max(insert_at - 1, 1)
                if reference_row and reference_row != insert_at:
                    _clone_row_style(reference_row, insert_at)
                # Shift tracked rows beneath the insertion point
                for info in employee_rows.values():
                    if info["row"] >= insert_at:
                        info["row"] += 1
                for other_block in section_blocks.values():
                    other_block["blank_rows"] = [
                        r + 1 if r >= insert_at else r for r in other_block["blank_rows"]
                    ]
                    end_row = other_block.get("end_row")
                    if end_row and end_row >= insert_at:
                        other_block["end_row"] = end_row + 1
                row_num = insert_at
                block["end_row"] = (block.get("end_row") or insert_at) + 1

            ws[f'D{row_num}'] = display_name.upper()
            block["template_row"] = block.get("template_row") or row_num
            employee_rows[(employee_key, section_name)] = {"row": row_num, "section": section_name}
            return row_num

        def _normalize_label_cell(value: Any) -> Optional[str]:
            if not isinstance(value, str):
                return None
            cleaned = "".join(ch for ch in value.upper() if ch.isalnum())
            return cleaned or None

        occupancy_row: Optional[int] = None
        crew_row_map: dict[str, int] = {}
        carrier_label_map = {carrier: _normalize_label_cell(carrier) for carrier in carriers}
        for row in range(1, ws.max_row + 1):
            normalized = _normalize_label_cell(ws[f'D{row}'].value)
            if not normalized:
                continue
            if normalized in {"OCC", "OCCUPANCY"} and occupancy_row is None:
                occupancy_row = row
                continue
            for carrier, carrier_norm in carrier_label_map.items():
                if not carrier_norm:
                    continue
                if normalized == carrier_norm or carrier_norm.startswith(normalized) or normalized.startswith(carrier_norm):
                    crew_row_map.setdefault(carrier, row)
                    break
        
        # Build map of employee -> primary section name (for cross-role tagging)
        emp_primary: dict[str, str] = {}
        for e in s.scalars(select(Employee)):
            sec = s.get(Section, e.section_id)
            if sec:
                emp_primary[e.name.upper()] = sec.name

        # Helpers to classify a raw assignment label into a section
        def shift_section_of(value: Optional[str]) -> Optional[str]:
            if not value or value in NEUTRAL_ASSIGNMENT_VALUES:
                return None
            if value in FRONT_DESK_SHIFTS:
                return "Front Desk"
            if value in BREAKFAST_SHIFTS:
                return "Breakfast Bar"
            if value in SHUTTLE_SHIFTS:
                return "Shuttle"
            if value in MAINTENANCE_SHIFTS:
                return "Maintenance"
            return None

        role_abbrev = {"Front Desk": "FD", "Breakfast Bar": "BB", "Shuttle": "SH", "Maintenance": "MA"}

        def _is_crew_shift_label(value: Optional[str], section_name: Optional[str] = None) -> bool:
            if section_name != "Shuttle":
                return False
            if not value or value in NEUTRAL_ASSIGNMENT_VALUES:
                return False
            if value == SHUTTLE_COMBO_LABEL:
                return True
            return _infer_shuttle_variant(value) == "Crew"

        # Fill in the shift data
        import re
        def _normalize_shift_display(raw: str) -> str:
            if not raw:
                return ""
            if raw == "Set":
                return "-"
            original = raw
            # Prefer time-only inside parentheses if present
            if "(" in raw and ")" in raw:
                start = raw.find("(") + 1
                end = raw.find(")", start)
                if end > start:
                    raw = raw[start:end]
            # Add spaces around en dash and hyphen
            raw = re.sub(r"\s*–\s*", " – ", raw)
            raw = re.sub(r"\s*-\s*", " - ", raw)
            # Lowercase AM/PM tokens (including attached to times)
            raw = re.sub(r"AM|PM", lambda m: m.group(0).lower(), raw)
            # Specific display tweak for the combo shuttle shift requested by the team
            if original.strip().lower() == SHUTTLE_COMBO_LABEL.lower():
                return "10:30am - 6:30pm (c)"
            return raw

        # Maps of employee -> set of date ISO strings
        vacation_days = ctx.get("vacation_days", {})
        dismissed_days = ctx.get("dismissed_days", {})
        for section_name in ["Breakfast Bar", "Front Desk", "Shuttle", "Maintenance"]:
            if section_name not in sections:
                continue
                
            section = sections[section_name]
            assignments = section["assignments"]
            
            for employee_name, employee_assignments in assignments.items():
                employee_key = employee_name.upper()
                row_num = _ensure_employee_row(section_name, employee_key, employee_name)
                if not row_num:
                    continue

                # Fill in the shifts for each day (columns E-K)
                for i, date_info in enumerate(dates):
                    col_letter = get_column_letter(5 + i)  # Start from column E (5)
                    date_key = date_info["key"]
                    date_label_md = date_info.get("label_md")
                    if not date_label_md:
                        try:
                            parsed_date = date.fromisoformat(date_key)
                            date_label_md = f"{parsed_date.month}/{parsed_date.day}"
                        except Exception:
                            date_label_md = ""
                    shift_value = employee_assignments[date_key]
                    cell = ws[f'{col_letter}{row_num}']

                    if shift_value is None or (isinstance(shift_value, str) and not shift_value.strip()):
                        cell.value = "-"
                        continue
                    if isinstance(shift_value, str) and shift_value.strip() == "-":
                        cell.value = "-"
                        continue

                    # For time-off, show as request type based on Vacation toggle
                    if shift_value in TIME_OFF_VALUES:
                        is_dismissed = date_key in (dismissed_days.get(employee_name, set()) or set())
                        is_vac = date_key in (vacation_days.get(employee_name, set()) or set())
                        if shift_value == REQ_VAC_LABEL or (is_dismissed and is_vac):
                            suffix = f" {date_label_md}" if date_label_md else ""
                            cell.value = f"REQ VAC{suffix}"
                        else:
                            suffix = f" {date_label_md}" if date_label_md else ""
                            cell.value = f"REQ OFF{suffix}"
                        continue

                    # Format shift display: time-only and normalized dash spacing
                    shift_display = _normalize_shift_display(shift_value)
                    # Front Desk export requirement: remove :00 and use hyphen
                    # Apply FD time formatting if the shift itself is a Front Desk shift,
                    # regardless of which section row we're writing into.
                    if shift_section_of(shift_value) == "Front Desk":
                        # collapse ":00am/pm" -> "am/pm"
                        shift_display = re.sub(r":00(?=(am|pm))", "", shift_display)
                        # use hyphen instead of en dash
                        shift_display = shift_display.replace("–", "-")
                        # normalize spacing around hyphen
                        shift_display = re.sub(r"\s*-\s*", " - ", shift_display)

                    # Set the cell value and apply crew fill when needed
                    cell.value = shift_display
                    if _is_crew_shift_label(shift_value, section_name):
                        cell.fill = CREW_EXCEL_FILL

        if occupancy_row is not None:
            for i, date_info in enumerate(dates):
                col_letter = get_column_letter(5 + i)
                cell = ws[f'{col_letter}{occupancy_row}']
                value = occupancy_values.get(date_info["key"])
                if value is None:
                    cell.value = None
                else:
                    # Always display occupancy with a trailing percent sign per export request
                    cell.value = f"{value}%"
                    cell.number_format = "General"

        if crew_row_map:
            for carrier, row_num in crew_row_map.items():
                per_day = carrier_arrivals.get(carrier) or {}
                for i, date_info in enumerate(dates):
                    col_letter = get_column_letter(5 + i)
                    cell = ws[f'{col_letter}{row_num}']
                    times = per_day.get(date_info["key"]) or []
                    if not times:
                        cell.value = None
                        continue
                    formatted_times = " / ".join(_format_aircrew_time_display(t) for t in sorted(times))
                    cell.value = formatted_times
                    existing_alignment = cell.alignment or Alignment()
                    cell.alignment = Alignment(
                        horizontal=existing_alignment.horizontal or "left",
                        vertical=existing_alignment.vertical or "top",
                        wrap_text=False,
                    )
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename like: "Sept 18 - Oct 15.xlsx"
        def month_abbrev(d: date) -> str:
            abbr = d.strftime('%b')
            # Prefer "Sept" over "Sep" for September
            return 'Sept' if d.month == 9 and abbr == 'Sep' else abbr
        end_date = week.start_date + timedelta(days=6)
        filename = f"{month_abbrev(week.start_date)} {week.start_date.day} - {month_abbrev(end_date)} {end_date.day}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )


if __name__ == "__main__":
    init_db_once()
    app.run(host="0.0.0.0", port=5003, debug=True)
